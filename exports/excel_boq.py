"""
Excel BOQ export — tender-grade, SANS 10142-1 compliant.

Sheets:
  1. Cover         — project info, contractor info, declarations
  2. Executive Summary — section subtotals + grand total
  3. Bill of Quantities — 14 sections, items with rates and totals
  4. Compliance    — SANS / NRS / SANS 10400-XA references + COC notice
  5. Acceptance    — payment terms, validity, signature block
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agent.shared import BillOfQuantities, BQSection, ContractorProfile, ProjectMetadata


# Brand palette (matches the blueprint UI: blueprint blue + ink + paper)
_INK = "0F1B3D"
_BLUEPRINT = "1E40AF"
_PAPER = "F5F2EA"
_PAPER_DARK = "EDEAE0"
_HAIRLINE = "C0BCB0"

_HEADER_FILL = PatternFill("solid", fgColor=_BLUEPRINT)
_SECTION_FILL = PatternFill("solid", fgColor=_INK)
_PAPER_FILL = PatternFill("solid", fgColor=_PAPER)
_PAPER_DARK_FILL = PatternFill("solid", fgColor=_PAPER_DARK)

_BORDER = Border(
    left=Side(style="thin", color=_HAIRLINE),
    right=Side(style="thin", color=_HAIRLINE),
    top=Side(style="thin", color=_HAIRLINE),
    bottom=Side(style="thin", color=_HAIRLINE),
)
_HEAVY_BOTTOM = Border(bottom=Side(style="medium", color=_INK))


def _money(cell) -> None:
    cell.number_format = '"R "#,##0.00'
    cell.alignment = Alignment(horizontal="right")


def _heading(cell, text: str, *, size: int = 18, color: str = _INK) -> None:
    cell.value = text
    cell.font = Font(name="Georgia", size=size, bold=True, color=color)


def _label(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(name="Calibri", size=10, bold=True, color=_INK)


def _value(cell, text) -> None:
    cell.value = text
    cell.font = Font(name="Calibri", size=10, color=_INK)


def export_boq_to_excel(
    boq: BillOfQuantities,
    *,
    project: Optional[ProjectMetadata] = None,
    contractor: Optional[ContractorProfile] = None,
    quote_ref: Optional[str] = None,
    validity_days: int = 30,
) -> bytes:
    """Generate a tender-grade .xlsx for the given BOQ."""
    project = project or ProjectMetadata(project_name=boq.project_name)
    contractor = contractor or ContractorProfile()
    issued = boq.generated_at
    valid_until = issued + timedelta(days=validity_days)
    quote_ref = quote_ref or f"AFP-{issued:%Y%m%d}-{boq.run_id[:6].upper()}"

    wb = Workbook()

    _build_cover(wb.active, boq, project, contractor, quote_ref, issued, valid_until)
    _build_executive_summary(wb.create_sheet("Executive Summary"), boq, contractor)
    _build_boq_sheet(wb.create_sheet("Bill of Quantities"), boq)
    _build_compliance(wb.create_sheet("Compliance"), boq, project)
    _build_acceptance(wb.create_sheet("Acceptance"), boq, contractor, valid_until)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Sheet 1 — Cover ──────────────────────────────────────────────────

def _build_cover(
    ws: Worksheet,
    boq: BillOfQuantities,
    project: ProjectMetadata,
    contractor: ContractorProfile,
    quote_ref: str,
    issued: datetime,
    valid_until: datetime,
) -> None:
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    # Column widths
    for c, w in [(1, 4), (2, 28), (3, 4), (4, 36)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    # Header band
    ws.row_dimensions[2].height = 40
    ws.merge_cells("B2:D2")
    cell = ws["B2"]
    _heading(cell, "BILL OF QUANTITIES", size=22, color=_BLUEPRINT)
    cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B3:D3")
    sub = ws["B3"]
    sub.value = "Tender Document · SANS 10142-1:2017"
    sub.font = Font(name="Calibri", size=10, italic=True, color=_INK)

    ws["B4"].border = _HEAVY_BOTTOM
    ws["C4"].border = _HEAVY_BOTTOM
    ws["D4"].border = _HEAVY_BOTTOM

    # Quote reference + dates (right-aligned column D)
    ws["D6"] = "Quote reference:"
    ws["D6"].font = Font(name="Calibri", size=9, color="6B7280")
    ws["D6"].alignment = Alignment(horizontal="right")
    ws["D7"] = quote_ref
    ws["D7"].font = Font(name="Georgia", size=14, bold=True, color=_INK)
    ws["D7"].alignment = Alignment(horizontal="right")
    ws["D8"] = f"Issued: {issued:%Y-%m-%d}    Valid until: {valid_until:%Y-%m-%d}"
    ws["D8"].font = Font(name="Calibri", size=9, color=_INK)
    ws["D8"].alignment = Alignment(horizontal="right")

    # Project block
    row = 11
    _heading(ws.cell(row=row, column=2), "PROJECT", size=12, color=_BLUEPRINT)
    row += 1
    for k, v in [
        ("Project name",     project.project_name or boq.project_name),
        ("Client",           project.client_name),
        ("Consultant",       project.consultant_name),
        ("Site address",     project.site_address),
        ("Drawing standard", project.standard or "SANS 10142-1:2017"),
        ("Pipeline used",    boq.pipeline.upper()),
    ]:
        _label(ws.cell(row=row, column=2), k)
        _value(ws.cell(row=row, column=4), v or "—")
        row += 1

    row += 1

    # Contractor block
    _heading(ws.cell(row=row, column=2), "CONTRACTOR", size=12, color=_BLUEPRINT)
    row += 1
    for k, v in [
        ("Company",            contractor.company_name),
        ("ECSA / CIDB number", contractor.registration_number),
        ("Contact person",     contractor.contact_name),
        ("Phone",              contractor.contact_phone),
        ("Email",              contractor.contact_email),
        ("VAT number",         contractor.vat_number),
        ("Physical address",   contractor.physical_address),
    ]:
        _label(ws.cell(row=row, column=2), k)
        _value(ws.cell(row=row, column=4), v or "—")
        row += 1

    row += 2

    # Statement
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    s = ws.cell(row=row, column=2)
    s.value = (
        "This Bill of Quantities is issued in accordance with SANS 10142-1:2017 "
        "(Wiring of Premises) and is subject to the SA standard tender conditions. "
        "All electrical work shall be certified by issue of a Certificate of "
        "Compliance (COC) on completion."
    )
    s.font = Font(name="Calibri", size=9.5, italic=True, color=_INK)
    s.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 60


# ─── Sheet 2 — Executive Summary ──────────────────────────────────────

def _build_executive_summary(
    ws: Worksheet,
    boq: BillOfQuantities,
    contractor: ContractorProfile,
) -> None:
    ws.sheet_view.showGridLines = False
    for c, w in [(1, 4), (2, 60), (3, 22)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:C2")
    _heading(ws["B2"], "EXECUTIVE SUMMARY", size=18, color=_BLUEPRINT)

    ws["B3"].border = _HEAVY_BOTTOM
    ws["C3"].border = _HEAVY_BOTTOM

    # Section subtotals
    row = 5
    by_section = boq.section_subtotals_zar
    if by_section:
        for section_value, subtotal in sorted(
            by_section.items(),
            key=lambda kv: _section_number_for(kv[0]),
        ):
            _value(ws.cell(row=row, column=2), section_value)
            cash = ws.cell(row=row, column=3, value=subtotal)
            cash.font = Font(name="Calibri", size=10, color=_INK)
            _money(cash)
            row += 1

    row += 1
    _section_total(ws, row, "Subtotal", boq.subtotal_zar, bold=True)
    row += 1
    _section_total(ws, row, f"Contingency ({boq.contingency_pct}%)", boq.contingency_zar)
    row += 1
    _section_total(ws, row, f"Markup ({boq.contractor_markup_pct}%)", boq.markup_zar)
    row += 1
    _section_total(ws, row, "Total excluding VAT", boq.total_excl_vat_zar, heavy=True)
    row += 1
    _section_total(ws, row, f"VAT ({boq.vat_pct}%)", boq.vat_zar)
    row += 1
    _section_total(ws, row, "TOTAL INCLUDING VAT", boq.total_incl_vat_zar, heavy=True, big=True)


def _section_total(
    ws: Worksheet,
    row: int,
    label: str,
    value: float,
    *,
    bold: bool = False,
    heavy: bool = False,
    big: bool = False,
) -> None:
    cell_label = ws.cell(row=row, column=2, value=label)
    cell_value = ws.cell(row=row, column=3, value=value)
    cell_label.alignment = Alignment(horizontal="right")
    cell_label.font = Font(
        name="Calibri",
        size=12 if big else 10,
        bold=bold or heavy or big,
        color=_INK,
    )
    cell_value.font = Font(
        name="Georgia" if big else "Calibri",
        size=14 if big else (11 if heavy else 10),
        bold=bold or heavy or big,
        color=_INK if not big else _BLUEPRINT,
    )
    _money(cell_value)
    if heavy or big:
        cell_label.border = _HEAVY_BOTTOM
        cell_value.border = _HEAVY_BOTTOM


def _section_number_for(section_value: str) -> int:
    """Extract the leading section number from a BQSection.value string."""
    try:
        return int(section_value.split(":", 1)[0].replace("SECTION", "").strip())
    except (ValueError, IndexError):
        return 99


# ─── Sheet 3 — Bill of Quantities ─────────────────────────────────────

def _build_boq_sheet(ws: Worksheet, boq: BillOfQuantities) -> None:
    ws.sheet_view.showGridLines = False
    headers = ["Item", "Description", "Unit", "Qty", "Unit Price (R)", "Total (R)", "Source", "Notes"]
    widths = {1: 10, 2: 56, 3: 8, 4: 10, 5: 16, 6: 18, 7: 14, 8: 30}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # Title row
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:H1")
    _heading(ws["A1"], "BILL OF QUANTITIES", size=14, color=_BLUEPRINT)

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = Font(name="Calibri", size=10, bold=True, color=_PAPER)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_cursor = 4
    by_section: dict[BQSection, list] = {}
    for it in boq.line_items:
        by_section.setdefault(it.section, []).append(it)

    for section in sorted(by_section.keys(), key=lambda s: s.section_number):
        # Section header
        sec_cell = ws.cell(row=row_cursor, column=1, value=section.value)
        sec_cell.fill = _SECTION_FILL
        sec_cell.font = Font(name="Calibri", size=10, bold=True, color=_PAPER)
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=8)
        ws.row_dimensions[row_cursor].height = 20
        row_cursor += 1

        section_total = 0.0
        for it in by_section[section]:
            ws.cell(row=row_cursor, column=1, value=it.item_number_str)
            ws.cell(row=row_cursor, column=2, value=it.description)
            ws.cell(row=row_cursor, column=3, value=it.unit)
            ws.cell(row=row_cursor, column=4, value=it.qty)
            unit = ws.cell(row=row_cursor, column=5, value=it.unit_price_zar)
            total = ws.cell(row=row_cursor, column=6, value=it.total_zar)
            ws.cell(row=row_cursor, column=7, value=it.source.value)
            ws.cell(row=row_cursor, column=8, value=it.notes)

            _money(unit)
            _money(total)
            for c in range(1, 9):
                ws.cell(row=row_cursor, column=c).border = _BORDER

            section_total += it.total_zar
            row_cursor += 1

        # Section subtotal
        sub_label = ws.cell(row=row_cursor, column=2, value="Section subtotal")
        sub_label.font = Font(name="Calibri", size=10, bold=True, italic=True, color=_INK)
        sub_label.alignment = Alignment(horizontal="right")
        sub_value = ws.cell(row=row_cursor, column=6, value=round(section_total, 2))
        sub_value.font = Font(name="Calibri", size=10, bold=True, color=_INK)
        sub_value.fill = _PAPER_DARK_FILL
        _money(sub_value)
        for c in range(1, 9):
            ws.cell(row=row_cursor, column=c).border = _BORDER
        row_cursor += 2

    # Grand totals block at the bottom
    _grand_total(ws, row_cursor, "Subtotal", boq.subtotal_zar)
    row_cursor += 1
    _grand_total(ws, row_cursor, f"Contingency ({boq.contingency_pct}%)", boq.contingency_zar)
    row_cursor += 1
    _grand_total(ws, row_cursor, f"Markup ({boq.contractor_markup_pct}%)", boq.markup_zar)
    row_cursor += 1
    _grand_total(ws, row_cursor, "Total excluding VAT", boq.total_excl_vat_zar, heavy=True)
    row_cursor += 1
    _grand_total(ws, row_cursor, f"VAT ({boq.vat_pct}%)", boq.vat_zar)
    row_cursor += 1
    _grand_total(ws, row_cursor, "TOTAL INCLUDING VAT", boq.total_incl_vat_zar, heavy=True, big=True)


def _grand_total(
    ws: Worksheet,
    row: int,
    label: str,
    value: float,
    *,
    heavy: bool = False,
    big: bool = False,
) -> None:
    cell_label = ws.cell(row=row, column=5, value=label)
    cell_label.alignment = Alignment(horizontal="right")
    cell_label.font = Font(
        name="Calibri",
        size=12 if big else 10,
        bold=heavy or big,
        color=_INK,
    )
    cell_value = ws.cell(row=row, column=6, value=value)
    _money(cell_value)
    cell_value.font = Font(
        name="Georgia" if big else "Calibri",
        size=13 if big else 11 if heavy else 10,
        bold=heavy or big,
        color=_BLUEPRINT if big else _INK,
    )
    if heavy or big:
        cell_label.border = _HEAVY_BOTTOM
        cell_value.border = _HEAVY_BOTTOM


# ─── Sheet 4 — Compliance ─────────────────────────────────────────────

def _build_compliance(
    ws: Worksheet,
    boq: BillOfQuantities,
    project: ProjectMetadata,
) -> None:
    ws.sheet_view.showGridLines = False
    for c, w in [(1, 4), (2, 30), (3, 60)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:C2")
    _heading(ws["B2"], "COMPLIANCE & STANDARDS", size=18, color=_BLUEPRINT)
    ws["B3"].border = _HEAVY_BOTTOM
    ws["C3"].border = _HEAVY_BOTTOM

    standards = [
        (
            "SANS 10142-1:2017",
            "Wiring of premises — Low-voltage installations. The master code "
            "for all installations under 1000 V AC. Dictates conductor sizing, "
            "earthing, protection, points-per-circuit, and inspection.",
        ),
        (
            "NRS 034",
            "After Diversity Maximum Demand for residential installations. "
            "Drives supply sizing and dictates ADMD-based load calculations.",
        ),
        (
            "SANS 10400-XA",
            "Energy efficiency in buildings. Sets lighting power density caps "
            "(LPD W/m²) by occupancy class for commercial works.",
        ),
        (
            "SANS 10139",
            "Fire detection and alarm systems for buildings. Applies wherever "
            "the BOQ includes fire detection or evacuation systems.",
        ),
        (
            "ECSA / OHS Act",
            "All electrical work to be carried out under the supervision of a "
            "registered Electrician (Wireman) per the Occupational Health & "
            "Safety Act, 1993, and certified by COC on completion.",
        ),
    ]

    row = 5
    _label(ws.cell(row=row, column=2), "Standard")
    _label(ws.cell(row=row, column=3), "Application")
    ws.cell(row=row, column=2).border = _BORDER
    ws.cell(row=row, column=3).border = _BORDER
    ws.cell(row=row, column=2).fill = _PAPER_DARK_FILL
    ws.cell(row=row, column=3).fill = _PAPER_DARK_FILL
    row += 1

    for std, desc in standards:
        c1 = ws.cell(row=row, column=2, value=std)
        c2 = ws.cell(row=row, column=3, value=desc)
        c1.font = Font(name="Calibri", size=10, bold=True, color=_INK)
        c2.font = Font(name="Calibri", size=10, color=_INK)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c1.border = _BORDER
        c2.border = _BORDER
        ws.row_dimensions[row].height = 42
        row += 1

    row += 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    statement = ws.cell(row=row, column=2)
    statement.value = (
        "The undersigned contractor warrants that all electrical work supplied "
        "and installed under this Bill of Quantities will comply with SANS "
        "10142-1:2017 in all material respects, and that a Certificate of "
        "Compliance (COC) will be issued on completion of the works."
    )
    statement.font = Font(name="Calibri", size=10, italic=True, color=_INK)
    statement.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 70


# ─── Sheet 5 — Acceptance ─────────────────────────────────────────────

def _build_acceptance(
    ws: Worksheet,
    boq: BillOfQuantities,
    contractor: ContractorProfile,
    valid_until: datetime,
) -> None:
    ws.sheet_view.showGridLines = False
    for c, w in [(1, 4), (2, 30), (3, 50)]:
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[2].height = 32
    ws.merge_cells("B2:C2")
    _heading(ws["B2"], "ACCEPTANCE & PAYMENT", size=18, color=_BLUEPRINT)
    ws["B3"].border = _HEAVY_BOTTOM
    ws["C3"].border = _HEAVY_BOTTOM

    row = 5
    _label(ws.cell(row=row, column=2), "Quotation valid until:")
    _value(ws.cell(row=row, column=3), valid_until.strftime("%Y-%m-%d"))
    row += 1
    _label(ws.cell(row=row, column=2), "Payment terms:")
    _value(ws.cell(row=row, column=3), _payment_terms_explanation(contractor.payment_terms))
    row += 1
    _label(ws.cell(row=row, column=2), "Total payable (incl VAT):")
    cell = ws.cell(row=row, column=3, value=boq.total_incl_vat_zar)
    cell.font = Font(name="Georgia", size=14, bold=True, color=_BLUEPRINT)
    _money(cell)
    row += 3

    # Signature blocks
    _heading(ws.cell(row=row, column=2), "ACCEPTED BY CLIENT", size=11, color=_INK)
    _heading(ws.cell(row=row, column=3), "ISSUED BY CONTRACTOR", size=11, color=_INK)
    row += 2
    for label in ("Name:", "Signature:", "Date:", "Company:"):
        _label(ws.cell(row=row, column=2), label)
        _label(ws.cell(row=row, column=3), label)
        ws.row_dimensions[row].height = 30
        ws.cell(row=row, column=2).border = Border(bottom=Side(style="thin", color=_HAIRLINE))
        ws.cell(row=row, column=3).border = Border(bottom=Side(style="thin", color=_HAIRLINE))
        row += 2

    # Contractor company auto-filled if available
    if contractor.company_name:
        ws.cell(row=row, column=3, value=contractor.company_name).font = Font(
            name="Calibri", size=10, italic=True, color="6B7280"
        )


def _payment_terms_explanation(terms: str) -> str:
    explanations = {
        "40/40/20": "40% deposit on order · 40% on materials delivery · 20% on completion",
        "50/30/20": "50% deposit on order · 30% on materials delivery · 20% on completion",
        "30/30/30/10": "30% deposit · 30% mid-progress · 30% on commissioning · 10% on COC",
    }
    return explanations.get(terms, terms or "Per separate agreement")
