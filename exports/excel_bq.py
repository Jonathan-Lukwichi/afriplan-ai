"""
AfriPlan Electrical v4.6 — Professional Excel BQ Export

Based on industry-standard BOQ format (Wedela Electrical BOQ reference).

Features:
- 4 worksheets: Cover, BOQ, Summary, Discrepancy Register
- 7 columns: Item, Description, Unit, Qty, Rate, Amount, Drawing Ref.
- 14 sections matching SA tender document conventions
- Hierarchical item numbering (1.1, 2.3, etc.)
- Multi-line DB descriptions with circuit schedules
- Aggregated fixtures with location lists
- Section subtotals with Excel formulas
- VAT calculations
- Professional notes section
"""

import io
from typing import List, Optional, Dict
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from agent.models import (
    PricingResult, BQLineItem, BQSection, ContractorProfile,
    ExtractionResult, ItemConfidence, ProjectMetadata, Discrepancy
)


# Brand colors
BRAND_CYAN = "00D4FF"
BRAND_DARK = "0A0E1A"
HEADER_GRAY = "1F2937"
SECTION_GRAY = "374151"
SUBSECTION_GRAY = "4B5563"
WARNING_YELLOW = "FEF3C7"
DISCREPANCY_PINK = "FCE7F3"


def export_professional_bq(
    pricing: PricingResult,
    extraction: ExtractionResult,
    project_name: str = "Electrical Installation",
    contractor: Optional[ContractorProfile] = None,
    include_estimates: bool = False,
) -> bytes:
    """
    Export professional 4-sheet Excel BOQ.

    Args:
        pricing: PricingResult with quantity_bq
        extraction: ExtractionResult for metadata and discrepancies
        project_name: Project name for header
        contractor: Contractor profile for letterhead
        include_estimates: If True, fill in estimated prices (default: False)

    Returns:
        Excel file as bytes
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export")

    wb = Workbook()

    # Create sheets in order
    ws_cover = wb.active
    ws_cover.title = "Cover"

    ws_boq = wb.create_sheet("BOQ")
    ws_summary = wb.create_sheet("Summary")

    # Only create discrepancy sheet if there are discrepancies
    discrepancies = extraction.discrepancies if hasattr(extraction, 'discrepancies') else []
    if discrepancies:
        ws_discrepancy = wb.create_sheet("Discrepancy Register")

    # Setup styles
    _setup_styles(wb)

    # Build metadata from extraction
    metadata = extraction.metadata

    # Generate sheets
    _create_cover_sheet(ws_cover, metadata, project_name)
    section_rows = _create_boq_sheet(ws_boq, pricing.quantity_bq, include_estimates)
    _create_summary_sheet(ws_summary, section_rows, project_name, metadata)

    if discrepancies:
        _create_discrepancy_sheet(ws_discrepancy, discrepancies)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_quantity_bq(
    pricing: PricingResult,
    project_name: str = "Project",
    contractor: Optional[ContractorProfile] = None,
) -> bytes:
    """
    Export quantity-only BQ to Excel (legacy compatibility).

    This is the PRIMARY deliverable — quantities only, no prices.
    Contractor fills in their own prices.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"

    _setup_styles(wb)

    # Header
    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "BILL OF QUANTITIES — ELECTRICAL INSTALLATION"
    ws[f"A{row}"].font = Font(name="Arial", size=16, bold=True, color=BRAND_CYAN)
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 2
    ws[f"A{row}"] = f"Project: {project_name}"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
    row += 1
    ws[f"A{row}"] = f"Reference: AP-{datetime.now().strftime('%Y%m%d%H%M')}"

    # Table header
    row += 2
    headers = ["Item", "Description", "Unit", "Qty", "Rate (ZAR)", "Amount (ZAR)", "Drawing Ref."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    _set_column_widths(ws)

    # Data rows
    row += 1
    current_section = None
    current_subsection = None

    for item in pricing.quantity_bq:
        # Section header
        if item.section != current_section:
            current_section = item.section
            current_subsection = None
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = item.section.value
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=SECTION_GRAY, fill_type="solid")
            row += 1

        # Subsection header
        if item.subsection and item.subsection != current_subsection:
            current_subsection = item.subsection
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = item.subsection
            ws[f"A{row}"].font = Font(bold=True, italic=True)
            ws[f"A{row}"].fill = PatternFill(start_color=SUBSECTION_GRAY, fill_type="solid")
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            row += 1

        # Item row with hierarchical numbering
        item_number = f"{item.section.section_number}.{item.item_no}"
        ws.cell(row=row, column=1, value=item_number)
        ws.cell(row=row, column=2, value=item.description).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=item.unit)
        ws.cell(row=row, column=4, value=item.qty)
        ws.cell(row=row, column=5, value="").number_format = "#,##0.00"
        ws.cell(row=row, column=6).value = f"=D{row}*E{row}"
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        ws.cell(row=row, column=7, value=item.drawing_ref)

        # Highlight estimated items
        if item.source == ItemConfidence.ESTIMATED:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=WARNING_YELLOW, fill_type="solid"
                )

        # Highlight discrepancy items
        if item.is_discrepancy:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=DISCREPANCY_PINK, fill_type="solid"
                )

        row += 1

    # Summary
    _add_summary_rows(ws, row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_estimated_bq(
    pricing: PricingResult,
    project_name: str = "Project",
    contractor: Optional[ContractorProfile] = None,
) -> bytes:
    """
    Export estimated BQ to Excel (legacy compatibility).
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimated BQ"

    _setup_styles(wb)

    # Header with warning
    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "ESTIMATED BILL OF QUANTITIES — FOR REFERENCE ONLY"
    ws[f"A{row}"].font = Font(name="Arial", size=16, bold=True, color="F59E0B")
    ws[f"A{row}"].alignment = Alignment(horizontal="center")

    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "PRICES ARE ESTIMATES ONLY — USE YOUR OWN SUPPLIER PRICES"
    ws[f"A{row}"].font = Font(bold=True, color="EF4444")
    ws[f"A{row}"].fill = PatternFill(start_color="FEE2E2", fill_type="solid")

    row += 2
    ws[f"A{row}"] = f"Project: {project_name}"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"

    # Table header
    row += 2
    headers = ["Item", "Description", "Unit", "Qty", "Est. Price (R)", "Est. Total (R)", "Drawing Ref."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _set_column_widths(ws)

    # Data rows
    row += 1
    current_section = None

    for item in pricing.estimated_bq:
        if item.section != current_section:
            current_section = item.section
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = item.section.value
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=SECTION_GRAY, fill_type="solid")
            row += 1

        item_number = f"{item.section.section_number}.{item.item_no}"
        ws.cell(row=row, column=1, value=item_number)
        ws.cell(row=row, column=2, value=item.description).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=item.unit)
        ws.cell(row=row, column=4, value=item.qty)
        ws.cell(row=row, column=5, value=item.unit_price_zar).number_format = "#,##0.00"
        ws.cell(row=row, column=6, value=item.total_zar).number_format = "#,##0.00"
        ws.cell(row=row, column=7, value=item.drawing_ref)
        row += 1

    # Summary with actual values
    row += 1
    summaries = [
        ("Subtotal:", pricing.estimate_subtotal_zar),
        ("Contingency (5%):", pricing.estimate_contingency_zar),
        ("Margin:", pricing.estimate_margin_zar),
        ("Total excl VAT:", pricing.estimate_total_excl_vat_zar),
        ("VAT (15%):", pricing.estimate_vat_zar),
        ("TOTAL incl VAT:", pricing.estimate_total_incl_vat_zar),
    ]

    for label, value in summaries:
        ws.cell(row=row, column=5, value=label).font = Font(bold=True)
        ws.cell(row=row, column=6, value=value).number_format = "R #,##0.00"
        if "TOTAL" in label:
            ws.cell(row=row, column=6).font = Font(bold=True, size=12)
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _create_cover_sheet(ws, metadata: ProjectMetadata, project_name: str):
    """Create professional cover sheet."""
    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "BILL OF QUANTITIES"
    ws["A1"].font = Font(name="Arial", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Project section
    ws["A3"] = "PROJECT:"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = project_name or metadata.project_name
    ws["A4"].font = Font(size=14, bold=True)
    ws["A5"] = metadata.description

    # Discipline
    ws["A7"] = "DISCIPLINE:"
    ws["A7"].font = Font(bold=True)
    ws["A8"] = "ELECTRICAL INSTALLATION"
    ws["A8"].font = Font(size=12)

    # Consultant
    ws["A10"] = "CONSULTANT:"
    ws["A10"].font = Font(bold=True)
    ws["A12"] = metadata.consultant_name or "To be confirmed"
    ws["A13"] = metadata.consultant_address
    ws["A14"] = f"Tel: {metadata.consultant_tel}" if metadata.consultant_tel else ""

    # Client
    ws["A16"] = "CLIENT:"
    ws["A16"].font = Font(bold=True)
    ws["A18"] = metadata.client_name or "To be confirmed"
    ws["A19"] = metadata.client_address

    # Standard
    ws["A21"] = f"STANDARD: {metadata.standard or 'SANS 10142-1 | OHS Act 85 of 1993'}"

    # Doc number
    revision = metadata.revision or "1"
    ws["A23"] = f"DOC. NUMBER: AP-BOQ-{datetime.now().strftime('%Y%m%d')} | REV: {revision}"

    # Drawing references
    if metadata.drawing_numbers:
        ws["A25"] = "Drawing Reference: " + ", ".join(metadata.drawing_numbers[:4])
        if len(metadata.drawing_numbers) > 4:
            ws["A26"] = ", ".join(metadata.drawing_numbers[4:8])

    # Set column width
    ws.column_dimensions["A"].width = 80


def _create_boq_sheet(ws, items: List[BQLineItem], include_estimates: bool = False) -> Dict[BQSection, int]:
    """
    Create main BOQ sheet with professional format.

    Returns dict mapping sections to their subtotal row numbers.
    """
    # Headers
    headers = ["Item", "Description", "Unit", "Qty", "Rate (ZAR)", "Amount (ZAR)", "Drawing Ref."]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_GRAY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style='thin', color='000000')
        )

    _set_column_widths(ws)

    row = 2
    current_section = None
    current_subsection = None
    section_start_rows: Dict[BQSection, int] = {}
    section_subtotal_rows: Dict[BQSection, int] = {}

    for item in items:
        # Section header
        if item.section != current_section:
            # Add subtotal for previous section
            if current_section is not None:
                section_subtotal_rows[current_section] = row
                ws.cell(row=row, column=1, value=f"{current_section.value} SUBTOTAL")
                ws.merge_cells(f"A{row}:E{row}")
                ws.cell(row=row, column=1).font = Font(bold=True)
                start = section_start_rows[current_section]
                ws.cell(row=row, column=6).value = f"=SUM(F{start}:F{row-1})"
                ws.cell(row=row, column=6).number_format = "#,##0.00"
                ws.cell(row=row, column=6).font = Font(bold=True)
                row += 2

            current_section = item.section
            current_subsection = None
            section_start_rows[current_section] = row + 1

            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = item.section.value
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", size=11)
            ws[f"A{row}"].fill = PatternFill(start_color=SECTION_GRAY, fill_type="solid")
            row += 1

        # Subsection header
        if item.subsection and item.subsection != current_subsection:
            current_subsection = item.subsection
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = item.subsection
            ws[f"A{row}"].font = Font(bold=True, italic=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill(start_color=SUBSECTION_GRAY, fill_type="solid")
            row += 1

        # Discrepancy warning row
        if item.is_discrepancy and item.discrepancy_note:
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"] = f"[!] {item.discrepancy_note}"
            ws[f"A{row}"].font = Font(italic=True, color="B91C1C")
            ws[f"A{row}"].fill = PatternFill(start_color=DISCREPANCY_PINK, fill_type="solid")
            row += 1

        # Item row
        item_number = f"{item.section.section_number}.{item.item_no}"
        ws.cell(row=row, column=1, value=item_number)

        desc_cell = ws.cell(row=row, column=2, value=item.description)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=row, column=3, value=item.unit)
        ws.cell(row=row, column=4, value=item.qty if item.qty != int(item.qty) else int(item.qty))

        # Rate column - empty for contractor or estimated
        if include_estimates and item.unit_price_zar > 0:
            ws.cell(row=row, column=5, value=item.unit_price_zar).number_format = "#,##0.00"
            ws.cell(row=row, column=6, value=item.total_zar).number_format = "#,##0.00"
        else:
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6).value = f"=D{row}*E{row}"
            ws.cell(row=row, column=6).number_format = "#,##0.00"

        ws.cell(row=row, column=7, value=item.drawing_ref)

        # Highlight based on source
        if item.source == ItemConfidence.ESTIMATED:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=WARNING_YELLOW, fill_type="solid"
                )

        if item.is_discrepancy:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=DISCREPANCY_PINK, fill_type="solid"
                )

        # Adjust row height for multi-line descriptions
        if "\n" in item.description:
            ws.row_dimensions[row].height = 15 * (item.description.count("\n") + 1)

        row += 1

    # Final section subtotal
    if current_section is not None:
        section_subtotal_rows[current_section] = row
        ws.cell(row=row, column=1, value=f"{current_section.value} SUBTOTAL")
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        start = section_start_rows[current_section]
        ws.cell(row=row, column=6).value = f"=SUM(F{start}:F{row-1})"
        ws.cell(row=row, column=6).number_format = "#,##0.00"
        ws.cell(row=row, column=6).font = Font(bold=True)
        row += 2

    # Grand totals
    row += 1
    ws.cell(row=row, column=5, value="SUBTOTAL (EXCL. VAT):").font = Font(bold=True)
    subtotal_refs = "+".join([f"F{r}" for r in section_subtotal_rows.values()])
    ws.cell(row=row, column=6).value = f"={subtotal_refs}"
    ws.cell(row=row, column=6).number_format = "#,##0.00"
    ws.cell(row=row, column=6).font = Font(bold=True)
    subtotal_row = row

    row += 1
    ws.cell(row=row, column=5, value="VAT (15%):").font = Font(bold=True)
    ws.cell(row=row, column=6).value = f"=F{subtotal_row}*0.15"
    ws.cell(row=row, column=6).number_format = "#,##0.00"
    vat_row = row

    row += 1
    ws.cell(row=row, column=5, value="GRAND TOTAL (INCL. VAT):").font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).value = f"=F{subtotal_row}+F{vat_row}"
    ws.cell(row=row, column=6).number_format = "#,##0.00"
    ws.cell(row=row, column=6).font = Font(bold=True, size=12)

    return section_subtotal_rows


def _create_summary_sheet(ws, section_rows: Dict[BQSection, int], project_name: str, metadata: ProjectMetadata):
    """Create summary sheet with section references."""
    ws["A1"] = "BOQ SUMMARY - ELECTRICAL INSTALLATION"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = project_name or metadata.project_name

    # Headers
    ws["A4"] = "Section"
    ws["B4"] = "Description"
    ws["C4"] = "Amount (ZAR)"
    for col in range(1, 4):
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill(start_color=HEADER_GRAY, fill_type="solid")
        ws.cell(row=4, column=col).font = Font(bold=True, color="FFFFFF")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18

    # Section subtotals
    row = 5
    for section in BQSection:
        section_num = section.section_number
        section_name = section.value.replace(f"SECTION {section_num}: ", "")

        ws.cell(row=row, column=1, value=section_num)
        ws.cell(row=row, column=2, value=section_name)

        if section in section_rows:
            ws.cell(row=row, column=3).value = f"=BOQ!F{section_rows[section]}"
        else:
            ws.cell(row=row, column=3).value = 0

        ws.cell(row=row, column=3).number_format = "#,##0.00"
        row += 1

    # Totals
    row += 1
    first_section_row = 5
    last_section_row = row - 2

    ws.cell(row=row, column=2, value="SUBTOTAL (EXCL. VAT)").font = Font(bold=True)
    ws.cell(row=row, column=3).value = f"=SUM(C{first_section_row}:C{last_section_row})"
    ws.cell(row=row, column=3).number_format = "#,##0.00"
    ws.cell(row=row, column=3).font = Font(bold=True)
    subtotal_row = row

    row += 1
    ws.cell(row=row, column=2, value="VAT (15%)").font = Font(bold=True)
    ws.cell(row=row, column=3).value = f"=C{subtotal_row}*0.15"
    ws.cell(row=row, column=3).number_format = "#,##0.00"
    vat_row = row

    row += 1
    ws.cell(row=row, column=2, value="GRAND TOTAL (INCL. VAT)").font = Font(bold=True, size=12)
    ws.cell(row=row, column=3).value = f"=C{subtotal_row}+C{vat_row}"
    ws.cell(row=row, column=3).number_format = "#,##0.00"
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)

    # Notes section
    row += 3
    ws.cell(row=row, column=1, value="Notes:").font = Font(bold=True)

    notes = [
        "1. All rates to be inclusive of labour, materials, and installation.",
        "2. Cable lengths include 10% allowance for routing/wastage.",
        "3. All work to comply with SANS 10142-1 and OHS Act 85 of 1993.",
        "4. Yellow highlighted items have estimated quantities - verify on site.",
        "5. Pink highlighted items derive from discrepancy resolution - CONFIRM WITH DESIGNER.",
        "6. Solar PV system (panels, inverters, batteries) excluded - electrical provisions only.",
        "7. HVAC units excluded - electrical isolator connections only.",
        "8. Contractor to verify all quantities on site prior to commencement.",
        "9. Rate column to be completed by tenderer.",
    ]

    for note in notes:
        row += 1
        ws.cell(row=row, column=1, value=note)


def _create_discrepancy_sheet(ws, discrepancies: List[Discrepancy]):
    """Create discrepancy register sheet."""
    ws["A1"] = "DRAWING DISCREPANCY REGISTER"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "SLD Schedules vs Floor Plan Layouts"

    # Headers
    headers = ["No.", "Distribution Board", "SLD Shows", "Floor Plan Shows",
               "Discrepancy", "Impact on BOQ", "Action Required"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=HEADER_GRAY, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 30

    # Data rows
    row = 5
    for i, disc in enumerate(discrepancies, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=disc.distribution_board).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3, value=disc.sld_shows).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4, value=disc.floor_plan_shows).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5, value=disc.discrepancy).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=6, value=disc.impact_on_boq).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=7, value=disc.action_required).alignment = Alignment(wrap_text=True)

        ws.row_dimensions[row].height = 45
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "RECOMMENDATION: Request designer to issue revised SLD with updated circuit schedules matching floor plan layouts before construction."
    ws[f"A{row}"].font = Font(bold=True, color="B91C1C")


def _set_column_widths(ws):
    """Set standard column widths for BOQ sheet."""
    ws.column_dimensions["A"].width = 8    # Item
    ws.column_dimensions["B"].width = 70   # Description
    ws.column_dimensions["C"].width = 8    # Unit
    ws.column_dimensions["D"].width = 8    # Qty
    ws.column_dimensions["E"].width = 14   # Rate
    ws.column_dimensions["F"].width = 14   # Amount
    ws.column_dimensions["G"].width = 25   # Drawing Ref


def _add_summary_rows(ws, start_row: int):
    """Add summary calculation rows at the end of BOQ."""
    row = start_row + 1

    ws.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=SUM(F1:F{start_row})").number_format = "#,##0.00"

    row += 1
    ws.cell(row=row, column=5, value="Contingency (5%):").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=F{row-1}*0.05").number_format = "#,##0.00"

    row += 1
    ws.cell(row=row, column=5, value="Total excl VAT:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=F{row-2}+F{row-1}").number_format = "#,##0.00"

    row += 1
    ws.cell(row=row, column=5, value="VAT (15%):").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"=F{row-1}*0.15").number_format = "#,##0.00"

    row += 1
    ws.cell(row=row, column=5, value="TOTAL incl VAT:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=6, value=f"=F{row-2}+F{row-1}")
    ws.cell(row=row, column=6).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).number_format = "#,##0.00"

    # Notes
    row += 2
    ws[f"A{row}"] = "Notes:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    ws[f"A{row}"] = "* Yellow highlighted items have estimated quantities - please verify"
    row += 1
    ws[f"A{row}"] = "* Pink highlighted items derive from discrepancy resolution - confirm with designer"
    row += 1
    ws[f"A{row}"] = "* All prices exclude VAT unless stated"
    row += 1
    ws[f"A{row}"] = "* Quotation valid for 30 days"


def _setup_styles(wb: Workbook) -> None:
    """Setup workbook styles."""
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = "R #,##0.00"

    try:
        wb.add_named_style(currency_style)
    except ValueError:
        pass  # Style already exists
