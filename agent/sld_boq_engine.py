"""
SLD → BOQ Engine v1.0
=====================
Simulates the 10-step electrical contractor process for building a
Bill of Quantities from PDF electrical drawings (layout + SLDs).

Pipeline:
  Step 1: Page classification (layout vs SLD vs schedule)
  Step 2: DB board hierarchy extraction from SLDs (AI Vision)
  Step 3: Breaker schedule extraction per DB
  Step 4: Cable specification extraction from SLDs
  Step 5: Cable route distance extraction from layout
  Step 6: Fixture counting from layout legend
  Step 7: Cross-page correlation (link SLD specs ↔ layout distances)
  Step 8: Containment and accessories derivation
  Step 9: Labour calculation
  Step 10: BOQ assembly with pricing

Author: JLWanalytics / AfriPlan Electrical v5.2
Standard: SANS 10142-1:2017
"""

import fitz  # PyMuPDF
import base64
import json
import re
import time
import io
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Any
from enum import Enum


# ─────────────────────────────────────────────────────────────────
# DATA MODELS
# ─────────────────────────────────────────────────────────────────

class PageType(Enum):
    LAYOUT = "layout"
    SLD = "sld"
    SCHEDULE = "schedule"
    LEGEND = "legend"
    UNKNOWN = "unknown"


@dataclass
class Breaker:
    """A single circuit breaker in a DB board."""
    ref: str           # e.g. "Q1", "Q12"
    rating_a: int      # Amps: 10, 20, 32, 63, 100, 150, 200, 250
    poles: int = 1     # 1 = SP, 3 = TP, 4 = TP+N
    type: str = "MCB"  # MCB, MCCB, ELCB, RCD
    feeds: str = ""    # What it feeds: "L1", "ISO3", "SPARE", "DB-PFA"
    cable_spec: str = ""   # e.g. "1.5mm² GP WIRE", "25mm² PVC SWA PVC"
    load_w: int = 0        # Load in watts if specified


@dataclass
class DBBoard:
    """A distribution board extracted from an SLD."""
    name: str              # e.g. "DB-CR", "DB-1", "DB-PFA"
    rating: str = ""       # e.g. "400V, 200A, 15kA, 50Hz, 3PH+N+E"
    main_switch_a: int = 0 # Main switch amps
    fed_from: str = ""     # Parent DB name
    incoming_cable: str = ""  # e.g. "95mm²x4core copper PVC SWA PVC"
    breakers: List[Breaker] = field(default_factory=list)
    total_ways: int = 0
    spare_ways: int = 0
    page_num: int = 0
    description: str = ""  # e.g. "Pool Block", "Community Hall"

    @property
    def used_ways(self) -> int:
        return len([b for b in self.breakers if b.feeds != "SPARE"])


@dataclass
class CableRun:
    """A cable run linking two points."""
    from_db: str
    to_db: str
    spec: str          # e.g. "95mm²x4core copper PVC SWA PVC"
    length_m: float = 0.0
    source: str = ""   # "sld" or "layout" or "cross-ref"


@dataclass
class LayoutFixture:
    """A fixture extracted from the layout legend."""
    description: str
    symbol_code: str = ""
    quantity: int = 0
    category: str = ""  # "lights", "switches", "sockets", "codes"


@dataclass
class BOQItem:
    """A single line item in the Bill of Quantities."""
    section: str       # "A: DB Boards", "B: Cables", etc.
    ref: str           # Item reference number
    description: str
    unit: str          # "each", "m", "roll", "hr"
    quantity: float
    rate: float        # Unit price in ZAR
    amount: float = 0  # quantity × rate
    source_page: str = ""  # Which page(s) this was derived from

    def __post_init__(self):
        self.amount = round(self.quantity * self.rate, 2)


@dataclass
class BOQResult:
    """Complete BOQ output."""
    project_name: str
    project_description: str
    consultant: str
    standard: str
    db_boards: List[DBBoard]
    cable_runs: List[CableRun]
    fixtures: List[LayoutFixture]
    boq_items: List[BOQItem]
    total_excl_vat: float = 0
    vat: float = 0
    total_incl_vat: float = 0
    contingency: float = 0
    grand_total: float = 0
    extraction_method: str = ""
    pages_processed: int = 0
    ai_cost_zar: float = 0
    processing_time_s: float = 0


# ─────────────────────────────────────────────────────────────────
# PRICING DATABASE (SA 2025 Commercial/Industrial rates)
# ─────────────────────────────────────────────────────────────────

# Extended pricing for commercial/industrial items not in residential constants
COMMERCIAL_DB_PRICING = {
    # DB Board enclosures (commercial grade)
    "db_12way_surface": {"desc": "DB Board 12-way surface mount", "price": 1200},
    "db_18way_surface": {"desc": "DB Board 18-way surface mount", "price": 1800},
    "db_24way_surface": {"desc": "DB Board 24-way surface mount", "price": 2500},
    "db_36way_surface": {"desc": "DB Board 36-way surface mount", "price": 3800},
    "db_48way_surface": {"desc": "DB Board 48-way surface mount", "price": 5200},

    # Main switches / MCCBs
    "mccb_100a": {"desc": "MCCB 100A 3P+N", "price": 2800},
    "mccb_150a": {"desc": "MCCB 150A 3P+N", "price": 3500},
    "mccb_200a": {"desc": "MCCB 200A 3P+N", "price": 4200},
    "mccb_250a": {"desc": "MCCB 250A 3P+N", "price": 5500},
    "mccb_400a": {"desc": "MCCB 400A 3P+N", "price": 8500},

    # MCBs (single pole)
    "mcb_10a_sp": {"desc": "MCB 10A single pole", "price": 65},
    "mcb_16a_sp": {"desc": "MCB 16A single pole", "price": 65},
    "mcb_20a_sp": {"desc": "MCB 20A single pole", "price": 70},
    "mcb_32a_sp": {"desc": "MCB 32A single pole", "price": 85},
    "mcb_40a_sp": {"desc": "MCB 40A single pole", "price": 95},
    "mcb_50a_sp": {"desc": "MCB 50A single pole", "price": 110},
    "mcb_63a_sp": {"desc": "MCB 63A single pole", "price": 130},

    # MCBs (3-pole)
    "mcb_10a_3p": {"desc": "MCB 10A 3-pole", "price": 180},
    "mcb_16a_3p": {"desc": "MCB 16A 3-pole", "price": 180},
    "mcb_20a_3p": {"desc": "MCB 20A 3-pole", "price": 195},
    "mcb_32a_3p": {"desc": "MCB 32A 3-pole", "price": 280},
    "mcb_40a_3p": {"desc": "MCB 40A 3-pole", "price": 320},
    "mcb_50a_3p": {"desc": "MCB 50A 3-pole", "price": 360},
    "mcb_63a_3p": {"desc": "MCB 63A 3-pole", "price": 420},

    # Earth leakage
    "elcb_63a_30ma": {"desc": "ELCB 63A 30mA 4-pole", "price": 950},
    "elcb_80a_30ma": {"desc": "ELCB 80A 30mA 4-pole", "price": 1150},
    "elcb_100a_30ma": {"desc": "ELCB 100A 30mA 4-pole", "price": 1450},

    # Contactors
    "contactor_25a": {"desc": "Contactor 25A 3-pole", "price": 850},
    "contactor_32a": {"desc": "Contactor 32A 3-pole", "price": 950},
    "contactor_40a": {"desc": "Contactor 40A 3-pole", "price": 1100},
    "contactor_63a": {"desc": "Contactor 63A 3-pole", "price": 1400},

    # Surge protection
    "spd_type2": {"desc": "Surge protection Type 2", "price": 1800},
    "spd_type1_2": {"desc": "Surge protection Type 1+2", "price": 3200},
}

COMMERCIAL_CABLE_PRICING = {
    # GP Wire (internal wiring) — per 100m
    "gp_1.5mm_100m": {"desc": "GP wire 1.5mm² (lighting)", "unit": "100m", "price": 450},
    "gp_2.5mm_100m": {"desc": "GP wire 2.5mm² (power)", "unit": "100m", "price": 750},
    "gp_4mm_100m": {"desc": "GP wire 4mm²", "unit": "100m", "price": 1200},
    "gp_6mm_100m": {"desc": "GP wire 6mm²", "unit": "100m", "price": 1800},

    # SWA Cable (external/underground) — per metre
    "swa_4mm_4c": {"desc": "4mm² 4-core PVC SWA PVC", "unit": "m", "price": 85},
    "swa_6mm_4c": {"desc": "6mm² 4-core PVC SWA PVC", "unit": "m", "price": 110},
    "swa_10mm_4c": {"desc": "10mm² 4-core PVC SWA PVC", "unit": "m", "price": 165},
    "swa_16mm_4c": {"desc": "16mm² 4-core PVC SWA PVC", "unit": "m", "price": 230},
    "swa_25mm_4c": {"desc": "25mm² 4-core PVC SWA PVC", "unit": "m", "price": 340},
    "swa_35mm_4c": {"desc": "35mm² 4-core PVC SWA PVC", "unit": "m", "price": 450},
    "swa_50mm_4c": {"desc": "50mm² 4-core PVC SWA PVC", "unit": "m", "price": 620},
    "swa_70mm_4c": {"desc": "70mm² 4-core PVC SWA PVC", "unit": "m", "price": 850},
    "swa_95mm_4c": {"desc": "95mm² 4-core PVC SWA PVC", "unit": "m", "price": 1150},
    "swa_120mm_4c": {"desc": "120mm² 4-core PVC SWA PVC", "unit": "m", "price": 1450},

    # Cable glands
    "gland_20mm": {"desc": "Cable gland 20mm (CW type)", "unit": "each", "price": 45},
    "gland_25mm": {"desc": "Cable gland 25mm (CW type)", "unit": "each", "price": 55},
    "gland_32mm": {"desc": "Cable gland 32mm (CW type)", "unit": "each", "price": 75},
    "gland_40mm": {"desc": "Cable gland 40mm (CW type)", "unit": "each", "price": 95},
    "gland_50mm": {"desc": "Cable gland 50mm (CW type)", "unit": "each", "price": 120},
    "gland_63mm": {"desc": "Cable gland 63mm (CW type)", "unit": "each", "price": 155},
    "gland_75mm": {"desc": "Cable gland 75mm (CW type)", "unit": "each", "price": 195},
}

LABOUR_RATES_COMMERCIAL = {
    "db_install_small": {"desc": "Install DB board (≤24 way)", "unit": "each", "price": 2500},
    "db_install_large": {"desc": "Install DB board (>24 way)", "unit": "each", "price": 4500},
    "swa_cable_trench": {"desc": "Trench & lay SWA cable", "unit": "m", "price": 180},
    "swa_cable_pull": {"desc": "Pull SWA cable (existing trench)", "unit": "m", "price": 85},
    "gp_wire_pull": {"desc": "Pull GP wire in conduit", "unit": "m", "price": 45},
    "breaker_install": {"desc": "Install & label breaker", "unit": "each", "price": 85},
    "contactor_install": {"desc": "Install contactor", "unit": "each", "price": 250},
    "light_point": {"desc": "Light point complete", "unit": "point", "price": 350},
    "power_point": {"desc": "Power point complete", "unit": "point", "price": 380},
    "motor_connection": {"desc": "Motor connection (pump/HVAC)", "unit": "each", "price": 1200},
    "testing_commissioning": {"desc": "Testing & commissioning per DB", "unit": "each", "price": 2500},
    "coc_commercial": {"desc": "COC inspection commercial", "unit": "each", "price": 3500},
    "electrician_rate": {"desc": "Qualified electrician per hour", "unit": "hr", "price": 420},
    "assistant_rate": {"desc": "Electrician assistant per hour", "unit": "hr", "price": 220},
    "earth_system": {"desc": "Earth system complete per DB", "unit": "each", "price": 1500},
}

CONTAINMENT_PRICING = {
    "trench_excavation": {"desc": "Trench excavation 600mm deep", "unit": "m", "price": 120},
    "trench_backfill": {"desc": "Trench backfill and compaction", "unit": "m", "price": 65},
    "warning_tape": {"desc": "Underground warning tape", "unit": "m", "price": 8},
    "sand_bedding": {"desc": "Cable sand bedding", "unit": "m", "price": 35},
    "cable_tray_100mm": {"desc": "Cable tray 100mm galvanized", "unit": "m", "price": 180},
    "cable_tray_200mm": {"desc": "Cable tray 200mm galvanized", "unit": "m", "price": 280},
    "cable_tray_300mm": {"desc": "Cable tray 300mm galvanized", "unit": "m", "price": 380},
    "conduit_25mm_heavy": {"desc": "Steel conduit 25mm heavy gauge", "unit": "m", "price": 85},
    "conduit_32mm_heavy": {"desc": "Steel conduit 32mm heavy gauge", "unit": "m", "price": 110},
    "cable_marker": {"desc": "Underground cable marker post", "unit": "each", "price": 250},
    "earth_rod_1500": {"desc": "Earth rod 1500mm copper clad", "unit": "each", "price": 280},
    "earth_bar_12way": {"desc": "Earth bar 12-way", "unit": "each", "price": 120},
}


# ─────────────────────────────────────────────────────────────────
# STEP 1: PAGE CLASSIFIER
# ─────────────────────────────────────────────────────────────────

def classify_page(page: fitz.Page, page_num: int) -> Tuple[PageType, str]:
    """Classify a PDF page as layout, SLD, schedule, or legend."""
    text = page.get_text("text").upper()

    # Check for SLD indicators
    sld_keywords = ["SLD", "SINGLE LINE", "400V", "3PH+N+E", "MCCB",
                     "INCOMING MAIN", "FED FROM"]
    layout_keywords = ["BRAAI", "HIKING", "OUTDOOR", "GATE", "PARKING",
                       "EXISTING MINI SUB", "POOL BLOCK", "GUARD"]

    sld_score = sum(1 for kw in sld_keywords if kw in text)
    layout_score = sum(1 for kw in layout_keywords if kw in text)

    # Also check page reference in title block area
    # SLD pages usually have "-SLD" in the drawing number
    if "-SLD" in text:
        return PageType.SLD, "SLD drawing"

    if layout_score > sld_score:
        return PageType.LAYOUT, "Site layout"

    if sld_score > 0:
        return PageType.SLD, "SLD drawing"

    # For pages with minimal text (images), assume SLD if not page 1
    if len(text.strip()) < 50 and page_num > 0:
        return PageType.SLD, "SLD drawing (image-based)"

    return PageType.LAYOUT if page_num == 0 else PageType.UNKNOWN, "Unknown"


# ─────────────────────────────────────────────────────────────────
# STEP 2-4: AI VISION SLD PARSER
# ─────────────────────────────────────────────────────────────────

SLD_EXTRACTION_PROMPT = """You are an expert South African electrical engineer analyzing a Single Line Diagram (SLD).

Extract ALL information from this SLD page into the following JSON structure.
Be meticulous — every breaker, every cable spec, every label matters for the Bill of Quantities.

Return ONLY valid JSON (no markdown, no explanation):

{
  "page_title": "drawing number from title block e.g. WD-KIOSK-01-SLD",
  "page_description": "what this page shows e.g. Kiosk Distribution",
  "part_description": "from title block PART: field",

  "db_boards": [
    {
      "name": "DB board name e.g. DB-CR, DB-1, DB-PFA",
      "rating": "full rating string e.g. 400V, 200A, 15kA, 50Hz, 3PH+N+E",
      "main_switch_a": 200,
      "fed_from": "parent DB name or EXISTING MINI SUB",
      "incoming_cable": "cable spec e.g. 95mm²x4core copper PVC SWA PVC cable",
      "description": "what area it serves",
      "breakers": [
        {
          "ref": "Q1",
          "rating_a": 63,
          "poles": 3,
          "type": "MCB or MCCB or ELCB",
          "feeds": "what it feeds: L1, ISO3, DB-PFA, SPARE etc",
          "cable_spec": "1.5mm² GP WIRE or 2.5mm² GP WIRE or 4mm² PVC SWA PVC etc",
          "load_w": 384
        }
      ]
    }
  ],

  "circuit_loads": [
    {
      "circuit": "L1",
      "description": "e.g. 8X48W lighting",
      "total_watts": 384,
      "feeds_to": "what area/equipment"
    }
  ],

  "cable_runs": [
    {
      "from": "DB-CR",
      "to": "DB-1",
      "spec": "50mm² 4-core PVC SWA PVC",
      "noted_length": "if any distance is noted"
    }
  ],

  "special_equipment": [
    {
      "type": "METER or HVAC or PUMP or CONTACTOR",
      "description": "details",
      "rating": "rating if shown"
    }
  ],

  "notes": ["any notes visible on the drawing"]
}

IMPORTANT RULES:
- Count EVERY breaker including spares (mark feeds as "SPARE")
- For breakers feeding lighting circuits (L1, L2...), record the wattage from the label below
- For breakers feeding isolators (ISO1, ISO2...), note what equipment they feed
- Cable specs are usually written along the cable lines (e.g. "1.5mm² GP WIRE", "2.5mm² GP WIRE")
- MCCB = Moulded Case Circuit Breaker (typically >100A)
- MCB = Miniature Circuit Breaker (typically ≤63A)
- If a contactor (square symbol with diagonal line) is shown, record it
- Record earth leakage devices (shown as a circle with arrow symbol)
- The title block is usually bottom-right with drawing number, consultant, contractor info
"""

LAYOUT_EXTRACTION_PROMPT = """You are an expert South African electrical engineer analyzing an electrical site layout drawing.

Extract ALL information from this layout page into the following JSON structure.
Focus on: DB board locations, cable routes with distances, fixture types and counts, and the legend.

Return ONLY valid JSON:

{
  "page_title": "drawing number from title block",
  "project_name": "project name from title block",
  "consultant": "consulting engineer name",
  "contractor": "contractor name if shown",
  "standard": "SANS standard referenced",

  "legend": {
    "switches": [
      {"symbol": "description", "code": "symbol code if any", "quantity_on_plan": 0}
    ],
    "power_sockets": [
      {"symbol": "description", "code": "symbol code if any", "quantity_on_plan": 0}
    ],
    "lights": [
      {"symbol": "description", "code": "symbol code if any", "quantity_on_plan": 0}
    ],
    "codes": [
      {"symbol": "description"}
    ]
  },

  "db_board_locations": [
    {
      "name": "DB-CR",
      "location_description": "near kiosk, central area",
      "fed_from": "existing mini sub"
    }
  ],

  "cable_routes": [
    {
      "from": "DB-CR",
      "to": "DB-1",
      "distance_m": 85,
      "route_description": "along main path"
    }
  ],

  "site_features": ["braai area", "hiking trail", "pool", "kiosk", "guard house"],

  "notes": ["OHS Act 85 of 1993", "Standard: SANS 10142-1"]
}

IMPORTANT:
- Read EVERY cable distance marked on the drawing (shown as e.g. "85m", "55m", "110m")
- List ALL DB boards shown on the plan and their relative positions
- Count fixtures if visible (light symbols, socket symbols)
- The legend/schedule at the top or side lists all fixture types with symbols
"""


def render_page_to_base64(page: fitz.Page, dpi: int = 200) -> str:
    """Render a PDF page to a base64 PNG image."""
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    return base64.b64encode(img_bytes).decode("utf-8")


def call_claude_vision(image_b64: str, prompt: str,
                       model: str = "claude-haiku-4-5-20241022",
                       api_key: str = "") -> Tuple[dict, float]:
    """Call Claude Vision API and return parsed JSON + cost estimate."""
    import anthropic

    client = anthropic.Anthropic(api_key=api_key) if api_key else anthropic.Anthropic()

    message = client.messages.create(
        model=model,
        max_tokens=8000,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/png",
                        "data": image_b64,
                    }
                },
                {
                    "type": "text",
                    "text": prompt,
                }
            ]
        }]
    )

    response_text = message.content[0].text

    # Estimate cost
    input_tokens = message.usage.input_tokens
    output_tokens = message.usage.output_tokens

    if "haiku" in model:
        cost_usd = (input_tokens * 0.80 + output_tokens * 4.0) / 1_000_000
    elif "sonnet" in model:
        cost_usd = (input_tokens * 3.0 + output_tokens * 15.0) / 1_000_000
    else:  # opus
        cost_usd = (input_tokens * 15.0 + output_tokens * 75.0) / 1_000_000

    cost_zar = cost_usd * 18.5  # Approximate USD→ZAR

    # Parse JSON from response
    parsed = _parse_json_response(response_text)

    return parsed, cost_zar


def _parse_json_response(text: str) -> dict:
    """Robustly parse JSON from AI response."""
    # Try direct parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Try extracting from markdown code block
    match = re.search(r'```(?:json)?\s*\n?(.*?)\n?```', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except json.JSONDecodeError:
            pass

    # Try finding JSON object
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        try:
            # Fix trailing commas
            cleaned = re.sub(r',\s*([}\]])', r'\1', match.group(0))
            return json.loads(cleaned)
        except json.JSONDecodeError:
            pass

    return {"error": "Failed to parse JSON", "raw": text[:500]}


# ─────────────────────────────────────────────────────────────────
# STEP 5: LAYOUT TEXT PARSER (Page 1 — has text layer)
# ─────────────────────────────────────────────────────────────────

def extract_layout_text(page: fitz.Page) -> dict:
    """Extract structured data from layout page that has a text layer."""
    blocks = page.get_text("dict")["blocks"]

    text_items = []
    for block in blocks:
        if "lines" not in block:
            continue
        for line in block["lines"]:
            for span in line["spans"]:
                t = span["text"].strip()
                if t and len(t) > 1:
                    text_items.append({
                        "text": t,
                        "x": span["bbox"][0],
                        "y": span["bbox"][1],
                        "size": span["size"],
                    })

    result = {
        "legend": {"switches": [], "power_sockets": [], "lights": [], "codes": []},
        "db_boards": [],
        "cable_distances": [],
        "notes": [],
    }

    # Extract legend items (typically at top of page, small font)
    for item in text_items:
        t = item["text"]

        # Switches
        if "switch" in t.lower() or "isolator" in t.lower():
            result["legend"]["switches"].append(t)

        # Sockets
        elif "socket" in t.lower():
            result["legend"]["power_sockets"].append(t)

        # Lights
        elif any(kw in t.lower() for kw in ["led", "light", "fluorescent", "downlight", "flood", "pole light", "vapor"]):
            result["legend"]["lights"].append(t)

        # DB board references
        elif re.match(r'DB[-\s]', t, re.IGNORECASE):
            result["db_boards"].append(t)

        # Cable distances (e.g., "85m", "55m")
        dist_match = re.search(r'(\d+)\s*m\b', t)
        if dist_match and item["size"] > 4:
            result["cable_distances"].append({
                "distance_m": int(dist_match.group(1)),
                "x": item["x"],
                "y": item["y"],
                "context": t
            })

        # Notes
        if "SANS" in t or "OHS" in t:
            result["notes"].append(t)

    return result


# ─────────────────────────────────────────────────────────────────
# STEP 6-7: CROSS-PAGE CORRELATOR
# ─────────────────────────────────────────────────────────────────

def correlate_sld_to_layout(sld_data: List[dict], layout_data: dict) -> List[CableRun]:
    """Cross-reference SLD cable specs with layout distances."""
    cable_runs = []

    # Build a lookup of known cable runs from SLDs
    sld_cables = {}
    for page_data in sld_data:
        if "error" in page_data:
            continue
        for db in page_data.get("db_boards", []):
            if db.get("fed_from") and db.get("incoming_cable"):
                key = (db["fed_from"], db["name"])
                sld_cables[key] = db["incoming_cable"]

        for cable in page_data.get("cable_runs", []):
            key = (cable["from"], cable["to"])
            sld_cables[key] = cable["spec"]

    # Match with layout distances
    layout_distances = {}
    for item in layout_data.get("cable_distances", []):
        layout_distances[item["distance_m"]] = item

    # Create cable runs
    for (from_db, to_db), spec in sld_cables.items():
        run = CableRun(
            from_db=from_db,
            to_db=to_db,
            spec=spec,
            source="sld"
        )
        cable_runs.append(run)

    return cable_runs


# ─────────────────────────────────────────────────────────────────
# STEP 8-10: BOQ GENERATOR
# ─────────────────────────────────────────────────────────────────

def _get_cable_price_per_m(spec: str) -> Tuple[str, float]:
    """Match a cable specification to a pricing database entry."""
    spec_upper = spec.upper()

    # Extract mm² size
    size_match = re.search(r'(\d+(?:\.\d+)?)\s*MM', spec_upper)
    if not size_match:
        return spec, 0

    size = float(size_match.group(1))

    if "SWA" in spec_upper or "ARMOURED" in spec_upper:
        # SWA cable
        size_map = {
            4: ("swa_4mm_4c", 85),
            6: ("swa_6mm_4c", 110),
            10: ("swa_10mm_4c", 165),
            16: ("swa_16mm_4c", 230),
            25: ("swa_25mm_4c", 340),
            35: ("swa_35mm_4c", 450),
            50: ("swa_50mm_4c", 620),
            70: ("swa_70mm_4c", 850),
            95: ("swa_95mm_4c", 1150),
            120: ("swa_120mm_4c", 1450),
        }
        closest = min(size_map.keys(), key=lambda x: abs(x - size))
        return COMMERCIAL_CABLE_PRICING[size_map[closest][0]]["desc"], size_map[closest][1]

    elif "GP" in spec_upper or "WIRE" in spec_upper:
        # GP wire — price per metre (convert from 100m rolls)
        size_map = {
            1.5: ("gp_1.5mm_100m", 4.50),
            2.5: ("gp_2.5mm_100m", 7.50),
            4: ("gp_4mm_100m", 12.00),
            6: ("gp_6mm_100m", 18.00),
        }
        closest = min(size_map.keys(), key=lambda x: abs(x - size))
        return COMMERCIAL_CABLE_PRICING[size_map[closest][0]]["desc"], size_map[closest][1]

    return spec, 0


def _get_breaker_price(rating_a: int, poles: int = 1, breaker_type: str = "MCB") -> Tuple[str, float]:
    """Get breaker price based on rating and type."""
    if breaker_type.upper() in ("MCCB",) or rating_a >= 100:
        mccb_map = {100: 2800, 150: 3500, 200: 4200, 250: 5500, 400: 8500}
        closest = min(mccb_map.keys(), key=lambda x: abs(x - rating_a))
        return f"MCCB {closest}A 3P+N", mccb_map[closest]

    if breaker_type.upper() == "ELCB":
        elcb_map = {63: 950, 80: 1150, 100: 1450}
        closest = min(elcb_map.keys(), key=lambda x: abs(x - rating_a))
        return f"ELCB {closest}A 30mA 4-pole", elcb_map[closest]

    if poles >= 3:
        mcb3_map = {10: 180, 16: 180, 20: 195, 32: 280, 40: 320, 50: 360, 63: 420}
        closest = min(mcb3_map.keys(), key=lambda x: abs(x - rating_a))
        return f"MCB {closest}A 3-pole", mcb3_map[closest]

    # Single pole MCB
    mcb_map = {10: 65, 16: 65, 20: 70, 32: 85, 40: 95, 50: 110, 63: 130}
    closest = min(mcb_map.keys(), key=lambda x: abs(x - rating_a))
    return f"MCB {closest}A single pole", mcb_map[closest]


def _get_db_enclosure_price(total_ways: int) -> Tuple[str, float]:
    """Get DB board enclosure price based on number of ways."""
    if total_ways <= 12:
        return "DB Board 12-way surface mount", 1200
    elif total_ways <= 18:
        return "DB Board 18-way surface mount", 1800
    elif total_ways <= 24:
        return "DB Board 24-way surface mount", 2500
    elif total_ways <= 36:
        return "DB Board 36-way surface mount", 3800
    else:
        return "DB Board 48-way surface mount", 5200


def _gland_size_for_cable(spec: str) -> Tuple[str, float]:
    """Determine cable gland size from cable spec."""
    size_match = re.search(r'(\d+(?:\.\d+)?)\s*MM', spec.upper())
    if not size_match:
        return "Cable gland 25mm (CW type)", 55

    size = float(size_match.group(1))
    if size <= 4:
        return "Cable gland 20mm (CW type)", 45
    elif size <= 10:
        return "Cable gland 25mm (CW type)", 55
    elif size <= 25:
        return "Cable gland 32mm (CW type)", 75
    elif size <= 50:
        return "Cable gland 40mm (CW type)", 95
    elif size <= 70:
        return "Cable gland 50mm (CW type)", 120
    elif size <= 95:
        return "Cable gland 63mm (CW type)", 155
    else:
        return "Cable gland 75mm (CW type)", 195


def generate_boq(db_boards: List[DBBoard], cable_runs: List[CableRun],
                 fixtures: List[LayoutFixture], project_info: dict) -> BOQResult:
    """Generate a complete priced BOQ from extracted data."""
    boq_items = []
    item_counter = {"A": 0, "B": 0, "C": 0, "D": 0, "E": 0, "F": 0, "G": 0, "H": 0}

    def add_item(section: str, desc: str, unit: str, qty: float, rate: float, source: str = ""):
        item_counter[section[0]] = item_counter.get(section[0], 0) + 1
        ref = f"{section[0]}{item_counter[section[0]]:02d}"
        boq_items.append(BOQItem(
            section=section, ref=ref, description=desc,
            unit=unit, quantity=qty, rate=rate, source_page=source
        ))

    # ─── SECTION A: PRELIMINARIES ───
    section = "A: Preliminaries & General"
    add_item(section, "Site establishment and mobilization", "sum", 1, 8500)
    add_item(section, "Health and safety file (OHS Act)", "sum", 1, 3500)
    add_item(section, "As-built drawings", "set", 1, 4500)
    add_item(section, "Project management and coordination", "sum", 1, 6000)

    # ─── SECTION B: DB BOARDS & SWITCHGEAR ───
    section = "B: DB Boards & Switchgear"
    for db in db_boards:
        # DB enclosure
        total_ways = max(len(db.breakers), db.total_ways, 12)
        enc_desc, enc_price = _get_db_enclosure_price(total_ways)
        add_item(section, f"{db.name} — {enc_desc}", "each", 1, enc_price, f"Page {db.page_num}")

        # Main switch
        if db.main_switch_a > 0:
            ms_desc, ms_price = _get_breaker_price(db.main_switch_a, 4,
                                                    "MCCB" if db.main_switch_a >= 100 else "MCB")
            add_item(section, f"{db.name} — Main switch: {ms_desc}", "each", 1, ms_price)

        # Earth leakage (mandatory per SANS 10142-1)
        elcb_rating = min(db.main_switch_a, 100) if db.main_switch_a > 0 else 63
        elcb_desc, elcb_price = _get_breaker_price(elcb_rating, 4, "ELCB")
        add_item(section, f"{db.name} — {elcb_desc}", "each", 1, elcb_price)

        # Surge protection (Type 2 per SANS 10142-1)
        add_item(section, f"{db.name} — Surge protection Type 2", "each", 1, 1800)

        # Individual breakers
        breaker_counts = {}
        for br in db.breakers:
            key = (br.rating_a, br.poles, br.type)
            breaker_counts[key] = breaker_counts.get(key, 0) + 1

        for (rating, poles, btype), count in sorted(breaker_counts.items()):
            br_desc, br_price = _get_breaker_price(rating, poles, btype)
            add_item(section, f"{db.name} — {br_desc}", "each", count, br_price)

        # Contactors (for isolator circuits with 30A contactors)
        contactor_count = len([b for b in db.breakers
                              if b.feeds.startswith("ISO") and b.feeds != "SPARE"])
        if contactor_count > 0:
            add_item(section, f"{db.name} — Contactor 30A 3-pole (for isolators)",
                    "each", contactor_count, 950)

        # Earth bar
        add_item(section, f"{db.name} — Earth bar 12-way", "each", 1, 120)

        # Busbar and DIN rail
        add_item(section, f"{db.name} — Busbar and DIN rail assembly", "set", 1, 650)

        # Labels
        add_item(section, f"{db.name} — Circuit labels and directory", "set", 1, 180)

    # ─── SECTION C: CABLES ───
    section = "C: Cable Supply"

    # SWA cables (external runs between DB boards)
    for run in cable_runs:
        if "SWA" in run.spec.upper() or run.length_m > 0:
            desc, price_per_m = _get_cable_price_per_m(run.spec)
            length = run.length_m if run.length_m > 0 else 50  # Estimate if unknown
            # Add 15% wastage
            length_with_waste = round(length * 1.15, 1)
            add_item(section, f"{run.from_db} → {run.to_db}: {desc} (incl. 15% waste)",
                    "m", length_with_waste, price_per_m,
                    f"SLD + Layout")

    # Internal wiring (GP wire per DB board)
    for db in db_boards:
        # Count circuits by cable type
        gp_15_count = len([b for b in db.breakers
                          if "1.5" in b.cable_spec or b.feeds.startswith("L")])
        gp_25_count = len([b for b in db.breakers
                          if "2.5" in b.cable_spec or b.feeds.startswith(("P", "ISO"))])

        if gp_15_count > 0:
            # Estimate 30m average run per lighting circuit
            total_m = gp_15_count * 30
            add_item(section, f"{db.name} — GP wire 1.5mm² (lighting circuits)",
                    "m", total_m, 4.50)

        if gp_25_count > 0:
            # Estimate 25m average run per power circuit
            total_m = gp_25_count * 25
            add_item(section, f"{db.name} — GP wire 2.5mm² (power circuits)",
                    "m", total_m, 7.50)

    # Earth wire
    total_earth_m = sum(max(r.length_m, 50) for r in cable_runs if "SWA" in r.spec.upper()) * 0.3
    if total_earth_m > 0:
        add_item(section, "Earth continuity conductor 10mm²", "m", round(total_earth_m), 12.00)

    # ─── SECTION D: CONTAINMENT ───
    section = "D: Containment & Trenching"

    # Trenching for external SWA runs
    total_trench_m = sum(max(r.length_m, 50) for r in cable_runs if "SWA" in r.spec.upper())
    if total_trench_m > 0:
        add_item(section, "Trench excavation 600mm deep", "m", round(total_trench_m), 120)
        add_item(section, "Cable sand bedding 100mm", "m", round(total_trench_m), 35)
        add_item(section, "Underground warning tape", "m", round(total_trench_m), 8)
        add_item(section, "Trench backfill and compaction", "m", round(total_trench_m), 65)

    # Cable glands (2 per SWA cable run)
    for run in cable_runs:
        if "SWA" in run.spec.upper():
            gland_desc, gland_price = _gland_size_for_cable(run.spec)
            add_item(section, f"{run.from_db}→{run.to_db}: {gland_desc}",
                    "each", 2, gland_price)

    # Cable markers
    swa_count = len([r for r in cable_runs if "SWA" in r.spec.upper()])
    if swa_count > 0:
        add_item(section, "Underground cable marker posts", "each", swa_count * 2, 250)

    # ─── SECTION E: FIXTURES & FITTINGS ───
    section = "E: Fixtures & Fittings"
    for fix in fixtures:
        if fix.quantity > 0:
            # Try to price from description
            price = _estimate_fixture_price(fix.description)
            add_item(section, fix.description, "each", fix.quantity, price)

    # ─── SECTION F: EARTHING & COMPLIANCE ───
    section = "F: Earthing & Compliance"
    for db in db_boards:
        add_item(section, f"{db.name} — Earth rod 1500mm copper clad", "each", 1, 280)
    add_item(section, "Earth resistance testing", "test", len(db_boards), 450)
    add_item(section, "Insulation resistance testing (all circuits)", "sum", 1, 2500)
    add_item(section, "Earth continuity testing", "sum", 1, 1500)

    # ─── SECTION G: LABOUR ───
    section = "G: Labour"
    for db in db_boards:
        ways = max(len(db.breakers), 12)
        labour_type = "large" if ways > 24 else "small"
        price = 4500 if labour_type == "large" else 2500
        add_item(section, f"Install {db.name} ({ways}-way DB board)", "each", 1, price)

    # SWA cable laying
    for run in cable_runs:
        if "SWA" in run.spec.upper():
            length = max(run.length_m, 50)
            add_item(section, f"Lay SWA cable {run.from_db}→{run.to_db}", "m", round(length), 180)

    # Internal wiring labour
    total_circuits = sum(len(db.breakers) for db in db_boards)
    add_item(section, "Pull GP wire and make off (all circuits)", "circuit", total_circuits, 350)

    # Motor connections
    motor_count = sum(1 for db in db_boards for b in db.breakers
                     if any(kw in b.feeds.upper() for kw in ["PUMP", "HEAT", "HVAC", "MOTOR"]))
    if motor_count > 0:
        add_item(section, "Motor connection and commissioning", "each", motor_count, 1200)

    # Testing and commissioning
    add_item(section, "Testing and commissioning per DB board", "each", len(db_boards), 2500)

    # COC
    add_item(section, "COC inspection and certificate (commercial)", "each", 1, 3500)

    # ─── SECTION H: PROVISIONAL ───
    section = "H: Provisional Sums"
    subtotal = sum(item.amount for item in boq_items)
    add_item(section, "Unforeseen works and site conditions (5%)", "sum", 1, round(subtotal * 0.05))
    add_item(section, "Price escalation allowance (3%)", "sum", 1, round(subtotal * 0.03))

    # ─── TOTALS ───
    total_excl = sum(item.amount for item in boq_items)
    vat = round(total_excl * 0.15, 2)

    return BOQResult(
        project_name=project_info.get("project_name", "Wedela Recreational Club"),
        project_description=project_info.get("description", "Electrical Upgrading"),
        consultant=project_info.get("consultant", "KABE Consulting Engineers"),
        standard=project_info.get("standard", "SANS 10142-1:2017"),
        db_boards=db_boards,
        cable_runs=cable_runs,
        fixtures=fixtures,
        boq_items=boq_items,
        total_excl_vat=total_excl,
        vat=vat,
        total_incl_vat=round(total_excl + vat, 2),
        contingency=round(subtotal * 0.05, 2),
        grand_total=round(total_excl + vat, 2),
    )


def _estimate_fixture_price(description: str) -> float:
    """Estimate fixture price from description."""
    desc_lower = description.lower()
    if "flood" in desc_lower:
        return 550 if "50w" in desc_lower else 350
    if "downlight" in desc_lower:
        return 120
    if "fluorescent" in desc_lower or "batten" in desc_lower:
        if "double" in desc_lower or "2x" in desc_lower:
            return 380
        return 280
    if "vapor" in desc_lower or "vapour" in desc_lower:
        return 380
    if "pole light" in desc_lower or "outdoor" in desc_lower:
        return 1200
    if "socket" in desc_lower:
        if "double" in desc_lower:
            return 95
        return 55
    if "switch" in desc_lower:
        if "isolator" in desc_lower:
            return 250
        if "day" in desc_lower or "night" in desc_lower:
            return 350
        if "2 lever" in desc_lower:
            return 65
        return 45
    return 250  # Default


# ─────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────

class SLDBOQEngine:
    """
    Main engine that orchestrates the 10-step BOQ building process.

    Usage:
        engine = SLDBOQEngine(api_key="sk-ant-...")
        result = engine.process_pdf(pdf_bytes)
    """

    def __init__(self, api_key: str = "", use_ai: bool = True,
                 sld_model: str = "claude-haiku-4-5-20241022",
                 layout_model: str = "claude-haiku-4-5-20241022"):
        self.api_key = api_key
        self.use_ai = use_ai
        self.sld_model = sld_model
        self.layout_model = layout_model
        self.total_cost_zar = 0.0
        self.stage_results = {}

    def process_pdf(self, pdf_bytes: bytes,
                    progress_callback=None) -> BOQResult:
        """
        Process a PDF electrical drawing set and generate a complete BOQ.

        Args:
            pdf_bytes: Raw PDF file bytes
            progress_callback: Optional callable(stage_num, stage_name, detail)

        Returns:
            BOQResult with complete priced BOQ
        """
        start_time = time.time()

        def report(stage: int, name: str, detail: str = ""):
            if progress_callback:
                progress_callback(stage, name, detail)

        # Open document
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_pages = len(doc)

        report(1, "Page Classification", f"Analyzing {total_pages} pages...")

        # ─── STEP 1: Classify pages ───
        page_types = {}
        for i in range(total_pages):
            page = doc[i]
            ptype, desc = classify_page(page, i)
            page_types[i] = (ptype, desc)

        self.stage_results["classification"] = page_types

        layout_pages = [i for i, (pt, _) in page_types.items() if pt == PageType.LAYOUT]
        sld_pages = [i for i, (pt, _) in page_types.items() if pt == PageType.SLD]

        report(2, "Layout Analysis", f"Processing {len(layout_pages)} layout page(s)...")

        # ─── STEP 2: Extract layout data (text layer) ───
        layout_data = {}
        for page_idx in layout_pages:
            page = doc[page_idx]
            layout_data = extract_layout_text(page)

            # Also use AI for layout if available
            if self.use_ai and self.api_key:
                report(2, "Layout AI Analysis", "Sending layout to AI Vision...")
                img_b64 = render_page_to_base64(page, dpi=150)
                ai_layout, cost = call_claude_vision(
                    img_b64, LAYOUT_EXTRACTION_PROMPT,
                    model=self.layout_model, api_key=self.api_key
                )
                self.total_cost_zar += cost
                layout_data["ai_analysis"] = ai_layout

        self.stage_results["layout"] = layout_data

        report(3, "SLD Extraction", f"Processing {len(sld_pages)} SLD page(s)...")

        # ─── STEP 3-4: Extract SLD data (AI Vision required) ───
        sld_results = []
        db_boards = []
        all_cable_runs = []

        for idx, page_idx in enumerate(sld_pages):
            page = doc[page_idx]

            if self.use_ai and self.api_key:
                report(3, "SLD AI Extraction",
                       f"Page {page_idx + 1}/{total_pages} — sending to {self.sld_model}...")
                img_b64 = render_page_to_base64(page, dpi=200)
                sld_data, cost = call_claude_vision(
                    img_b64, SLD_EXTRACTION_PROMPT,
                    model=self.sld_model, api_key=self.api_key
                )
                self.total_cost_zar += cost
                sld_results.append(sld_data)

                # Parse DB boards from AI response
                for db_data in sld_data.get("db_boards", []):
                    breakers = []
                    for br in db_data.get("breakers", []):
                        breakers.append(Breaker(
                            ref=br.get("ref", ""),
                            rating_a=br.get("rating_a", 0),
                            poles=br.get("poles", 1),
                            type=br.get("type", "MCB"),
                            feeds=br.get("feeds", ""),
                            cable_spec=br.get("cable_spec", ""),
                            load_w=br.get("load_w", 0),
                        ))

                    db = DBBoard(
                        name=db_data.get("name", f"DB-{page_idx}"),
                        rating=db_data.get("rating", ""),
                        main_switch_a=db_data.get("main_switch_a", 0),
                        fed_from=db_data.get("fed_from", ""),
                        incoming_cable=db_data.get("incoming_cable", ""),
                        breakers=breakers,
                        total_ways=len(breakers),
                        spare_ways=len([b for b in breakers if b.feeds == "SPARE"]),
                        page_num=page_idx + 1,
                        description=db_data.get("description", ""),
                    )
                    db_boards.append(db)

                # Parse cable runs
                for cable in sld_data.get("cable_runs", []):
                    all_cable_runs.append(CableRun(
                        from_db=cable.get("from", ""),
                        to_db=cable.get("to", ""),
                        spec=cable.get("spec", ""),
                        length_m=float(cable.get("noted_length", "0").replace("m", "").strip() or 0),
                        source="sld"
                    ))
            else:
                # No AI — create placeholder
                sld_results.append({"note": "AI disabled — manual entry required"})

        self.stage_results["sld_extraction"] = sld_results

        report(5, "Cross-Reference", "Linking SLD specs to layout distances...")

        # ─── STEP 5-7: Cross-reference ───
        # Add cable runs from DB board incoming cables
        for db in db_boards:
            if db.fed_from and db.incoming_cable:
                # Check if this run already exists
                exists = any(r.from_db == db.fed_from and r.to_db == db.name
                           for r in all_cable_runs)
                if not exists:
                    all_cable_runs.append(CableRun(
                        from_db=db.fed_from,
                        to_db=db.name,
                        spec=db.incoming_cable,
                        source="sld"
                    ))

        # Try to match distances from layout
        if layout_data.get("ai_analysis"):
            ai_routes = layout_data["ai_analysis"].get("cable_routes", [])
            for route in ai_routes:
                for run in all_cable_runs:
                    from_match = (route.get("from", "").upper() in run.from_db.upper() or
                                 run.from_db.upper() in route.get("from", "").upper())
                    to_match = (route.get("to", "").upper() in run.to_db.upper() or
                               run.to_db.upper() in route.get("to", "").upper())
                    if from_match and to_match and run.length_m == 0:
                        run.length_m = float(route.get("distance_m", 0))
                        run.source = "cross-ref"

        self.stage_results["cable_runs"] = all_cable_runs

        report(6, "Fixture Counting", "Extracting fixtures from legend...")

        # ─── STEP 6: Fixture extraction ───
        fixtures = []
        legend_items = layout_data.get("legend", {})
        for category, items in legend_items.items():
            for item_text in items:
                fixtures.append(LayoutFixture(
                    description=item_text,
                    category=category,
                    quantity=0,  # Need AI or manual count
                ))

        # Use AI layout analysis for fixture counts if available
        if layout_data.get("ai_analysis"):
            ai_legend = layout_data["ai_analysis"].get("legend", {})
            for category in ["switches", "power_sockets", "lights"]:
                for ai_item in ai_legend.get(category, []):
                    for fix in fixtures:
                        if (ai_item.get("symbol", "").lower() in fix.description.lower() or
                            fix.description.lower() in ai_item.get("symbol", "").lower()):
                            fix.quantity = ai_item.get("quantity_on_plan", 0)

        self.stage_results["fixtures"] = fixtures

        report(8, "BOQ Generation", "Building priced Bill of Quantities...")

        # ─── STEPS 8-10: Generate BOQ ───
        project_info = {
            "project_name": "The Upgrading of Wedela Recreational Club",
            "description": "Electrical Upgrading — Outside Lights, Kiosk, Pool Block, "
                          "Community Hall, Pool Pumps, Heat Pumps",
            "consultant": "KABE Consulting Engineers",
            "contractor": "Choma-Malanga Engineering",
            "standard": "SANS 10142-1:2017",
        }

        # Try to get project info from AI layout analysis
        if layout_data.get("ai_analysis"):
            ai = layout_data["ai_analysis"]
            if ai.get("project_name"):
                project_info["project_name"] = ai["project_name"]
            if ai.get("consultant"):
                project_info["consultant"] = ai["consultant"]

        boq_result = generate_boq(db_boards, all_cable_runs, fixtures, project_info)
        boq_result.extraction_method = f"AI Vision ({self.sld_model})" if self.use_ai else "Manual"
        boq_result.pages_processed = total_pages
        boq_result.ai_cost_zar = self.total_cost_zar
        boq_result.processing_time_s = round(time.time() - start_time, 1)

        doc.close()

        report(10, "Complete", f"BOQ generated: {len(boq_result.boq_items)} items, "
               f"R{boq_result.total_incl_vat:,.2f} incl. VAT")

        return boq_result

    def process_pdf_no_ai(self, pdf_bytes: bytes,
                          manual_db_boards: List[dict] = None) -> BOQResult:
        """
        Process without AI — uses text extraction + manual DB board data.
        Useful for testing or when no API key is available.
        """
        self.use_ai = False
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")

        # Extract what we can from text
        layout_data = {}
        for i in range(len(doc)):
            page = doc[i]
            text = page.get_text("text")
            if len(text) > 100:  # Likely the layout page
                layout_data = extract_layout_text(page)
                break

        # Use manual DB board data if provided
        db_boards = []
        cable_runs = []

        if manual_db_boards:
            for mdb in manual_db_boards:
                breakers = []
                for br in mdb.get("breakers", []):
                    breakers.append(Breaker(**br))
                db = DBBoard(
                    name=mdb["name"],
                    rating=mdb.get("rating", ""),
                    main_switch_a=mdb.get("main_switch_a", 0),
                    fed_from=mdb.get("fed_from", ""),
                    incoming_cable=mdb.get("incoming_cable", ""),
                    breakers=breakers,
                    total_ways=len(breakers),
                    page_num=mdb.get("page_num", 0),
                    description=mdb.get("description", ""),
                )
                db_boards.append(db)

                if db.fed_from and db.incoming_cable:
                    cable_runs.append(CableRun(
                        from_db=db.fed_from, to_db=db.name,
                        spec=db.incoming_cable,
                        length_m=mdb.get("cable_length_m", 0),
                        source="manual"
                    ))

        fixtures = []
        for cat, items in layout_data.get("legend", {}).items():
            for desc in items:
                fixtures.append(LayoutFixture(description=desc, category=cat))

        doc.close()

        project_info = {
            "project_name": "Wedela Recreational Club",
            "description": "Electrical Upgrading",
            "consultant": "KABE Consulting Engineers",
            "standard": "SANS 10142-1:2017",
        }

        return generate_boq(db_boards, cable_runs, fixtures, project_info)


# ─────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────

def export_boq_to_excel(boq: BOQResult) -> bytes:
    """Export BOQ to a professional Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ─── COVER SHEET ───
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.column_dimensions['A'].width = 25
    ws_cover.column_dimensions['B'].width = 50

    cover_data = [
        ("BILL OF QUANTITIES", ""),
        ("", ""),
        ("Project:", boq.project_name),
        ("Description:", boq.project_description),
        ("Consultant:", boq.consultant),
        ("Standard:", boq.standard),
        ("", ""),
        ("Total excl. VAT:", f"R {boq.total_excl_vat:,.2f}"),
        ("VAT (15%):", f"R {boq.vat:,.2f}"),
        ("TOTAL incl. VAT:", f"R {boq.total_incl_vat:,.2f}"),
        ("", ""),
        ("Pages Processed:", str(boq.pages_processed)),
        ("DB Boards Found:", str(len(boq.db_boards))),
        ("BOQ Line Items:", str(len(boq.boq_items))),
        ("Extraction Method:", boq.extraction_method),
        ("AI Cost:", f"R {boq.ai_cost_zar:,.2f}"),
        ("Processing Time:", f"{boq.processing_time_s}s"),
    ]

    header_font = Font(name='Calibri', size=16, bold=True, color='00D4FF')
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    total_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')

    for row_idx, (label, value) in enumerate(cover_data, 1):
        ws_cover.cell(row=row_idx, column=1, value=label).font = header_font if row_idx == 1 else label_font
        ws_cover.cell(row=row_idx, column=2, value=value).font = value_font
        if "TOTAL incl" in label:
            ws_cover.cell(row=row_idx, column=1).font = total_font
            ws_cover.cell(row=row_idx, column=1).fill = total_fill
            ws_cover.cell(row=row_idx, column=2).font = total_font
            ws_cover.cell(row=row_idx, column=2).fill = total_fill

    # ─── BOQ SHEET ───
    ws_boq = wb.create_sheet("Bill of Quantities")
    headers = ["Ref", "Section", "Description", "Unit", "Qty", "Rate (R)", "Amount (R)"]
    col_widths = [8, 30, 55, 8, 8, 15, 18]

    for i, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_boq.cell(row=1, column=i, value=header)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        ws_boq.column_dimensions[chr(64 + i)].width = width

    current_section = ""
    row = 2
    section_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    section_totals = {}

    for item in boq.boq_items:
        # Section header
        if item.section != current_section:
            current_section = item.section
            cell = ws_boq.cell(row=row, column=1, value="")
            row += 1
            for col in range(1, 8):
                cell = ws_boq.cell(row=row, column=col)
                cell.fill = section_fill
                cell.font = Font(name='Calibri', size=11, bold=True)
            ws_boq.cell(row=row, column=2, value=current_section)
            row += 1

        # Item row
        ws_boq.cell(row=row, column=1, value=item.ref).border = thin_border
        ws_boq.cell(row=row, column=2, value=item.section).border = thin_border
        ws_boq.cell(row=row, column=3, value=item.description).border = thin_border
        ws_boq.cell(row=row, column=4, value=item.unit).border = thin_border

        qty_cell = ws_boq.cell(row=row, column=5, value=item.quantity)
        qty_cell.border = thin_border
        qty_cell.number_format = '#,##0.0'

        rate_cell = ws_boq.cell(row=row, column=6, value=item.rate)
        rate_cell.border = thin_border
        rate_cell.number_format = 'R #,##0.00'

        amt_cell = ws_boq.cell(row=row, column=7, value=item.amount)
        amt_cell.border = thin_border
        amt_cell.number_format = 'R #,##0.00'

        # Track section totals
        section_totals[item.section] = section_totals.get(item.section, 0) + item.amount

        row += 1

    # Section subtotals
    row += 1
    ws_boq.cell(row=row, column=3, value="SECTION SUBTOTALS").font = Font(bold=True, size=12)
    row += 1
    for section, total in section_totals.items():
        ws_boq.cell(row=row, column=3, value=section).font = Font(bold=True)
        cell = ws_boq.cell(row=row, column=7, value=total)
        cell.number_format = 'R #,##0.00'
        cell.font = Font(bold=True)
        row += 1

    # Grand totals
    row += 1
    for label, value in [
        ("TOTAL EXCL. VAT", boq.total_excl_vat),
        ("VAT (15%)", boq.vat),
        ("TOTAL INCL. VAT", boq.total_incl_vat),
    ]:
        ws_boq.cell(row=row, column=3, value=label).font = Font(bold=True, size=12)
        cell = ws_boq.cell(row=row, column=7, value=value)
        cell.number_format = 'R #,##0.00'
        cell.font = Font(bold=True, size=12)
        if "INCL" in label:
            for col in range(3, 8):
                ws_boq.cell(row=row, column=col).fill = total_fill
                ws_boq.cell(row=row, column=col).font = Font(bold=True, size=12, color='FFFFFF')
        row += 1

    # ─── DB BOARDS SHEET ───
    ws_db = wb.create_sheet("DB Board Schedule")
    db_headers = ["DB Name", "Rating", "Main Switch", "Fed From", "Incoming Cable",
                  "Total Ways", "Used Ways", "Spare Ways", "Page"]
    for i, h in enumerate(db_headers, 1):
        cell = ws_db.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        ws_db.column_dimensions[chr(64 + i)].width = 20

    for row_idx, db in enumerate(boq.db_boards, 2):
        ws_db.cell(row=row_idx, column=1, value=db.name)
        ws_db.cell(row=row_idx, column=2, value=db.rating)
        ws_db.cell(row=row_idx, column=3, value=f"{db.main_switch_a}A")
        ws_db.cell(row=row_idx, column=4, value=db.fed_from)
        ws_db.cell(row=row_idx, column=5, value=db.incoming_cable)
        ws_db.cell(row=row_idx, column=6, value=db.total_ways)
        ws_db.cell(row=row_idx, column=7, value=db.used_ways)
        ws_db.cell(row=row_idx, column=8, value=db.spare_ways)
        ws_db.cell(row=row_idx, column=9, value=db.page_num)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────
# CONVENIENCE FUNCTIONS
# ─────────────────────────────────────────────────────────────────

def process_electrical_drawings(pdf_path: str, api_key: str = "",
                                use_ai: bool = True) -> BOQResult:
    """One-liner convenience function."""
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    engine = SLDBOQEngine(api_key=api_key, use_ai=use_ai)
    return engine.process_pdf(pdf_bytes)


def print_boq_summary(boq: BOQResult):
    """Print a formatted BOQ summary to console."""
    print(f"\n{'='*70}")
    print(f"  BILL OF QUANTITIES — {boq.project_name}")
    print(f"  {boq.project_description}")
    print(f"  Consultant: {boq.consultant}")
    print(f"  Standard: {boq.standard}")
    print(f"{'='*70}")

    print(f"\n  DB Boards: {len(boq.db_boards)}")
    for db in boq.db_boards:
        print(f"    {db.name} ({db.main_switch_a}A) ← {db.fed_from} "
              f"[{db.total_ways} ways, {db.spare_ways} spare]")

    print(f"\n  Cable Runs: {len(boq.cable_runs)}")
    for run in boq.cable_runs:
        dist = f"{run.length_m}m" if run.length_m > 0 else "TBM"
        print(f"    {run.from_db} → {run.to_db}: {run.spec} ({dist})")

    current_section = ""
    section_totals = {}
    for item in boq.boq_items:
        if item.section != current_section:
            if current_section:
                print(f"    {'─'*50}")
                print(f"    Section Total: R {section_totals.get(current_section, 0):,.2f}")
            current_section = item.section
            print(f"\n  ── {current_section} ──")

        print(f"    {item.ref} | {item.description[:45]:<45} | "
              f"{item.quantity:>6.0f} {item.unit:<5} × R{item.rate:>10,.2f} = R{item.amount:>12,.2f}")
        section_totals[item.section] = section_totals.get(item.section, 0) + item.amount

    if current_section:
        print(f"    {'─'*50}")
        print(f"    Section Total: R {section_totals.get(current_section, 0):,.2f}")

    print(f"\n{'='*70}")
    print(f"  TOTAL EXCL. VAT:  R {boq.total_excl_vat:>15,.2f}")
    print(f"  VAT (15%):        R {boq.vat:>15,.2f}")
    print(f"  TOTAL INCL. VAT:  R {boq.total_incl_vat:>15,.2f}")
    print(f"{'='*70}")
    print(f"  AI Cost: R{boq.ai_cost_zar:.2f} | Pages: {boq.pages_processed} | "
          f"Time: {boq.processing_time_s}s | Items: {len(boq.boq_items)}")
