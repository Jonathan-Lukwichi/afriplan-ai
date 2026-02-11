"""
AfriPlan Electrical - Excel Export Functions
Export BQ and calculations to professional Excel format
"""

import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_bq_to_excel(
    bq_items: list,
    project_info: dict = None,
    calculation_data: dict = None
) -> bytes:
    """
    Export Bill of Quantities to professional Excel format.

    Args:
        bq_items: List of BQ items with category, item, qty, unit, rate, total
        project_info: Optional project details for cover sheet
        calculation_data: Optional calculation results to include

    Returns:
        bytes: Excel file as bytes for download
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
    subheader_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    money_format = 'R #,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ─────────────────────────────────────────────
    # Sheet 1: Cover Page
    # ─────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover"

    # Title
    ws_cover.merge_cells('A1:F1')
    ws_cover['A1'] = "ELECTRICAL QUOTATION"
    ws_cover['A1'].font = Font(bold=True, size=24, color="00D4FF")
    ws_cover['A1'].alignment = Alignment(horizontal='center')

    ws_cover.merge_cells('A2:F2')
    ws_cover['A2'] = "AfriPlan Electrical - SANS 10142 Compliant"
    ws_cover['A2'].font = Font(size=14, italic=True)
    ws_cover['A2'].alignment = Alignment(horizontal='center')

    row = 4

    # Project Info
    if project_info:
        ws_cover[f'A{row}'] = "PROJECT DETAILS"
        ws_cover[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        for key, value in project_info.items():
            ws_cover[f'A{row}'] = key
            ws_cover[f'A{row}'].font = Font(bold=True)
            ws_cover[f'B{row}'] = str(value)
            row += 1

        row += 1

    # Summary
    ws_cover[f'A{row}'] = "QUOTATION SUMMARY"
    ws_cover[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    subtotal = sum(item["total"] for item in bq_items)
    vat = subtotal * 0.15
    total = subtotal + vat

    summary_items = [
        ("Subtotal (excl VAT)", subtotal),
        ("VAT (15%)", vat),
        ("TOTAL (incl VAT)", total),
    ]

    for label, value in summary_items:
        ws_cover[f'A{row}'] = label
        ws_cover[f'A{row}'].font = Font(bold=True)
        ws_cover[f'B{row}'] = value
        ws_cover[f'B{row}'].number_format = money_format
        row += 1

    ws_cover[f'A{row}'].font = Font(bold=True, size=12)
    ws_cover[f'B{row}'].font = Font(bold=True, size=12)

    # Date
    row += 2
    ws_cover[f'A{row}'] = "Date Generated:"
    ws_cover[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Column widths
    ws_cover.column_dimensions['A'].width = 25
    ws_cover.column_dimensions['B'].width = 30

    # ─────────────────────────────────────────────
    # Sheet 2: Bill of Quantities
    # ─────────────────────────────────────────────
    ws_bq = wb.create_sheet("Bill of Quantities")

    # Headers
    headers = ["Category", "Item Description", "Quantity", "Unit", "Rate (R)", "Total (R)"]
    for col, header in enumerate(headers, 1):
        cell = ws_bq.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Group by category
    categories = {}
    for item in bq_items:
        cat = item["category"]
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)

    row = 2
    for cat_name, items in categories.items():
        # Category header
        ws_bq.merge_cells(f'A{row}:F{row}')
        cell = ws_bq.cell(row=row, column=1, value=cat_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = subheader_fill
        cell.border = thin_border
        row += 1

        # Items in category
        for item in items:
            ws_bq.cell(row=row, column=1, value=item["category"]).border = thin_border
            ws_bq.cell(row=row, column=2, value=item["item"]).border = thin_border
            ws_bq.cell(row=row, column=3, value=item["qty"]).border = thin_border
            ws_bq.cell(row=row, column=4, value=item["unit"]).border = thin_border

            rate_cell = ws_bq.cell(row=row, column=5, value=item["rate"])
            rate_cell.number_format = money_format
            rate_cell.border = thin_border

            total_cell = ws_bq.cell(row=row, column=6, value=item["total"])
            total_cell.number_format = money_format
            total_cell.border = thin_border

            row += 1

        # Category subtotal
        cat_total = sum(i["total"] for i in items)
        ws_bq.cell(row=row, column=5, value="Subtotal:").font = Font(bold=True)
        ws_bq.cell(row=row, column=5).border = thin_border
        total_cell = ws_bq.cell(row=row, column=6, value=cat_total)
        total_cell.number_format = money_format
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        row += 2

    # Grand totals
    row += 1
    ws_bq.cell(row=row, column=5, value="Subtotal (excl VAT):").font = Font(bold=True)
    ws_bq.cell(row=row, column=6, value=subtotal).number_format = money_format

    row += 1
    ws_bq.cell(row=row, column=5, value="VAT (15%):").font = Font(bold=True)
    ws_bq.cell(row=row, column=6, value=vat).number_format = money_format

    row += 1
    ws_bq.cell(row=row, column=5, value="TOTAL (incl VAT):").font = Font(bold=True, size=14)
    total_cell = ws_bq.cell(row=row, column=6, value=total)
    total_cell.number_format = money_format
    total_cell.font = Font(bold=True, size=14)

    # Column widths
    ws_bq.column_dimensions['A'].width = 20
    ws_bq.column_dimensions['B'].width = 40
    ws_bq.column_dimensions['C'].width = 12
    ws_bq.column_dimensions['D'].width = 12
    ws_bq.column_dimensions['E'].width = 15
    ws_bq.column_dimensions['F'].width = 15

    # ─────────────────────────────────────────────
    # Sheet 3: Calculation Backup (if provided)
    # ─────────────────────────────────────────────
    if calculation_data:
        ws_calc = wb.create_sheet("Calculations")

        ws_calc['A1'] = "CALCULATION BACKUP"
        ws_calc['A1'].font = Font(bold=True, size=16)

        row = 3
        for key, value in calculation_data.items():
            if not key.startswith('_') and not callable(value):
                ws_calc.cell(row=row, column=1, value=str(key).replace('_', ' ').title())
                ws_calc.cell(row=row, column=1).font = Font(bold=True)

                if isinstance(value, (int, float)):
                    if 'cost' in key.lower() or 'price' in key.lower() or 'total' in key.lower():
                        ws_calc.cell(row=row, column=2, value=value).number_format = money_format
                    else:
                        ws_calc.cell(row=row, column=2, value=value)
                elif isinstance(value, list):
                    ws_calc.cell(row=row, column=2, value=str(len(value)) + " items")
                elif isinstance(value, dict):
                    ws_calc.cell(row=row, column=2, value=str(value)[:50] + "...")
                else:
                    ws_calc.cell(row=row, column=2, value=str(value))
                row += 1

        ws_calc.column_dimensions['A'].width = 25
        ws_calc.column_dimensions['B'].width = 40

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_load_study_to_excel(
    load_data: dict,
    project_name: str = "Load Study"
) -> bytes:
    """
    Export load study calculations to Excel.

    Args:
        load_data: Dictionary with load calculation results
        project_name: Name for the project

    Returns:
        bytes: Excel file as bytes for download
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Study"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"LOAD STUDY - {project_name}"
    ws['A1'].font = Font(bold=True, size=18)

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    row = 4

    # Headers
    headers = ["Parameter", "Value", "Unit", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Data
    for key, value in load_data.items():
        if not key.startswith('_') and not isinstance(value, (list, dict)):
            ws.cell(row=row, column=1, value=str(key).replace('_', ' ').title())

            if isinstance(value, float):
                ws.cell(row=row, column=2, value=round(value, 2))
            else:
                ws.cell(row=row, column=2, value=value)

            # Add units based on key name
            if 'kva' in key.lower():
                ws.cell(row=row, column=3, value='kVA')
            elif 'kw' in key.lower() or 'load' in key.lower():
                ws.cell(row=row, column=3, value='kW')
            elif 'percent' in key.lower():
                ws.cell(row=row, column=3, value='%')
            elif 'current' in key.lower() or '_a' in key.lower():
                ws.cell(row=row, column=3, value='A')
            elif 'voltage' in key.lower() or '_v' in key.lower():
                ws.cell(row=row, column=3, value='V')

            row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
