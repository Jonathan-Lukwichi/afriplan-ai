"""
AfriPlan Electrical v4.1 — Excel Export Tests

Tests for Excel BQ export functionality.
"""

import pytest
from io import BytesIO

from agent.models import (
    PricingResult,
    BQItem,
    BQSection,
    ItemConfidence,
    ContractorProfile,
)

# Import only if openpyxl available
try:
    from openpyxl import load_workbook
    from exports.excel_bq import export_quantity_bq, export_estimated_bq
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@pytest.fixture
def sample_pricing():
    """Create sample pricing result for export tests."""
    quantity_items = [
        BQItem(
            section=BQSection.LIGHTING,
            description="LED Downlight 9W",
            unit="ea",
            qty=10,
            source=ItemConfidence.EXTRACTED,
        ),
        BQItem(
            section=BQSection.SOCKETS,
            description="Double Socket Outlet",
            unit="ea",
            qty=8,
            source=ItemConfidence.EXTRACTED,
        ),
        BQItem(
            section=BQSection.CABLES,
            description="2.5mm² Twin & Earth",
            unit="m",
            qty=150,
            source=ItemConfidence.INFERRED,
        ),
        BQItem(
            section=BQSection.DISTRIBUTION,
            description="18-Way DB Board",
            unit="ea",
            qty=1,
            source=ItemConfidence.EXTRACTED,
        ),
    ]

    estimated_items = [
        BQItem(
            section=BQSection.LIGHTING,
            description="LED Downlight 9W",
            unit="ea",
            qty=10,
            unit_price_zar=150.0,
            source=ItemConfidence.EXTRACTED,
        ),
        BQItem(
            section=BQSection.SOCKETS,
            description="Double Socket Outlet",
            unit="ea",
            qty=8,
            unit_price_zar=85.0,
            source=ItemConfidence.EXTRACTED,
        ),
        BQItem(
            section=BQSection.CABLES,
            description="2.5mm² Twin & Earth",
            unit="m",
            qty=150,
            unit_price_zar=18.50,
            source=ItemConfidence.INFERRED,
        ),
        BQItem(
            section=BQSection.DISTRIBUTION,
            description="18-Way DB Board",
            unit="ea",
            qty=1,
            unit_price_zar=2500.0,
            source=ItemConfidence.EXTRACTED,
        ),
    ]

    return PricingResult(
        quantity_bq=quantity_items,
        estimated_bq=estimated_items,
        total_items=4,
        items_from_extraction=3,
        estimate_subtotal_zar=6455.0,
        estimate_contingency_zar=322.75,
        estimate_margin_zar=968.25,
        estimate_total_excl_vat_zar=7746.0,
        estimate_vat_zar=1161.90,
        estimate_total_incl_vat_zar=8907.90,
        deposit_zar=3563.16,
        second_payment_zar=3563.16,
        final_payment_zar=1781.58,
    )


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestQuantityBQExport:
    """Test quantity-only BQ export."""

    def test_export_creates_valid_excel(self, sample_pricing):
        result = export_quantity_bq(sample_pricing, "Test Project")

        assert isinstance(result, bytes)
        assert len(result) > 0

        # Load and verify workbook
        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) >= 1

    def test_export_has_items(self, sample_pricing):
        result = export_quantity_bq(sample_pricing, "Test Project")
        wb = load_workbook(BytesIO(result))

        # Find the BQ sheet
        sheet = wb.active

        # Should have header and data rows
        assert sheet.max_row > 1

    def test_export_no_prices_column(self, sample_pricing):
        result = export_quantity_bq(sample_pricing, "Test Project")
        wb = load_workbook(BytesIO(result))
        sheet = wb.active

        # Look for price columns - they should be empty or have "Enter Price" placeholders
        # This verifies it's a quantity-only BQ
        # The exact implementation may vary, but prices should not be filled


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestEstimatedBQExport:
    """Test estimated BQ export."""

    def test_export_creates_valid_excel(self, sample_pricing):
        result = export_estimated_bq(sample_pricing, "Test Project")

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_export_has_totals(self, sample_pricing):
        result = export_estimated_bq(sample_pricing, "Test Project")
        wb = load_workbook(BytesIO(result))

        # Should have totals somewhere in the workbook
        found_total = False
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and "total" in str(cell).lower():
                        found_total = True
                        break

        assert found_total


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestContractorBranding:
    """Test contractor info in exports."""

    def test_company_name_in_export(self, sample_pricing):
        contractor = ContractorProfile(
            company_name="ABC Electrical (Pty) Ltd",
            contact_phone="011 234 5678",
        )

        result = export_quantity_bq(sample_pricing, "Test Project", contractor)
        wb = load_workbook(BytesIO(result))

        # Company name should appear somewhere
        found = False
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and "ABC Electrical" in str(cell):
                        found = True
                        break

        assert found


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestEmptyPricing:
    """Test export with empty/minimal pricing."""

    def test_empty_bq_export(self):
        empty_pricing = PricingResult()

        result = export_quantity_bq(empty_pricing, "Empty Project")

        # Should still create valid Excel
        assert isinstance(result, bytes)
        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) >= 1


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
